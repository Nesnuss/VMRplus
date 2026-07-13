#!/usr/bin/env python3

import os
import sys
import pandas as pd
import argparse
from Bio import Entrez
from Bio import SeqIO
from argparse import RawTextHelpFormatter
import time
from io import StringIO
import subprocess
import http.client 
import urllib.error
import socket
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import re
from collections import defaultdict
import csv
import configparser
import shutil
import threading
import fcntl
from concurrent.futures import ThreadPoolExecutor, as_completed

# Global socket timeout (seconds) applied to ALL NCBI Entrez/urllib requests.
# If a request stalls beyond this, Python raises socket.timeout, which the
# existing retry loops now catch and retry. 120 s = 2 minutes.
socket.setdefaulttimeout(120)


version = "1.7.14"

help_text = """
VMR Program version """+version+""" -  jul 2026

Generate an incremented table of the ICTV's VMR/MSL table

(c) 2026. Rafael Santos da Silva & Arthur Gruber

Usage: VMR.py -i <tabela VMR> -o <tabela output>

-i input<VMR table>         VMR/MSL table
-o output <file name>      	Output table file
-s <sheet number>           Worksheet number in VMR/MSL table
-t <auxiliary table>        Terms table
-ts <sheet number>          Worksheet number in terms table
-c / -config <file.ini>     Configuration file (optional)
--generate-config           Generate a template config file and exit
-gb <yes/no>                When 'yes', downloads the full nucleotide genome
                            (FASTA) of every individual in the VMR/MSL table
                            into a 'genome_data/' directory, organised by
                            family. Files are named VMR<7-digit counter>_
                            <accession(s)>.fasta. Default: no.
-thread / -thr <N>          Number of parallel worker threads for NCBI
                            network calls (Entrez + refdb + cds_prot).
                            Requires an NCBI API key. With API key the
                            NCBI limit is 10 req/s; keep N <= 10 to stay
                            safely under that ceiling. Omitting this flag
                            runs the program sequentially (default).

Note: -c/--config cannot be combined with the parameters it covers
      (-i, -o, -s, -t, -ts). Use one or the other.
"""

# Parameters covered by the config file (conflict detection)
CONFIG_COVERED_PARAMS = {'-i', '-o', '-s', '-t', '-ts'}

# Default Tabajara parameters
TABAJARA_CON_DEFAULTS = {
    't': '0.5',
    'p': '50',
    'w': '15',
    'b': '20',
    'cs': 'yes',
    'm': 'c',
}

TABAJARA_DIS_DEFAULTS = {
    't': '0.5',
    'p': '50',
    'w': '15',
    'b': '20',
    'cs': 'yes',
    'm': 'd',
}

TEMPLATE_CONFIG = """\
; VMR Program - Configuration File Template
; Version: {version}
;
; Usage: VMR.py -c VMR_config.ini
;
; Lines starting with ';' are comments.
; Remove the leading ';' to activate a parameter.
; Parameters left blank will use the program's built-in defaults.

[general]
; Path to the VMR/MSL input table (.xlsx)
input =

; Output directory name
output = output_dir

; Worksheet number in the VMR/MSL table (1-based)
sheet = 1

; Path to the auxiliary terms table (.xlsx)
terms =

; Worksheet number in the terms table (1-based)
terms_sheet = 1

; Entrez email (required by NCBI)
email = rafass2003@gmail.com

; Entrez API key
api_key = 511ef882e71fdff1e01eaaa3177e47c43e09

; Download every genome (FASTA) from the VMR/MSL table into 'genome_data/'
; organised by family (yes/no). Default: no.
gb = no


[tabajara_con]
; Parameters for the CONSERVATIVE Tabajara run (tabajara.pl -m c)
; Any key written here is passed directly as -key value to tabajara.pl
; Add any extra tabajara parameter freely — unknown keys are forwarded as-is.
;
; Mode (should remain 'c' for the conservative run)
m = c
; Threshold
t = 0.5
; Percentage of sequences
p = 50
; Window size
w = 15
; Block size
b = 20
; Conserved sites (yes/no)
cs = yes


[tabajara_dis]
; Parameters for the DISCRIMINATORY Tabajara run (tabajara.pl -m d)
; Any key written here is passed directly as -key value to tabajara.pl
; Add any extra tabajara parameter freely — unknown keys are forwarded as-is.
;
; Mode (should remain 'd' for the discriminatory run)
m = d
; Threshold
t = 0.5
; Percentage of sequences
p = 50
; Window size
w = 15
; Block size
b = 20
; Conserved sites (yes/no)
cs = yes
"""


def generate_config_template():
    """Generates a VMR_config_template.ini file in the current directory."""
    template_path = "VMR_config_template.ini"
    content = TEMPLATE_CONFIG.format(version=version)
    with open(template_path, "w") as f:
        f.write(content)
    print(f"Config template generated: {template_path}")
    print("Edit the file and run: VMR.py -c VMR_config_template.ini")

def load_config(config_path):
    """
    Loads and validates a .ini config file.

    Returns a dict with keys:
        general       -> dict of [general] section
        tabajara_con  -> dict of [tabajara_con] section
        tabajara_dis  -> dict of [tabajara_dis] section

    Unknown keys in [general] generate a WARNING and are ignored.
    Unknown keys in [tabajara_*] are forwarded as-is to tabajara.pl.
    """

    KNOWN_GENERAL_KEYS = {'input', 'output', 'sheet', 'terms', 'terms_sheet', 'email', 'api_key', 'gb'}

    if not os.path.exists(config_path):
        print(f"Error: Config file not found: {config_path}")
        sys.exit(1)

    parser_cfg = configparser.ConfigParser(
        allow_no_value=False,
        inline_comment_prefixes=(';', '#')
    )
    parser_cfg.read(config_path, encoding='utf-8')

    config = {
        'general': {},
        'tabajara_con': dict(TABAJARA_CON_DEFAULTS),
        'tabajara_dis': dict(TABAJARA_DIS_DEFAULTS),
    }

    # ── [general] ──────────────────────────────────────────────────────────
    if parser_cfg.has_section('general'):
        for key, value in parser_cfg.items('general'):
            if key not in KNOWN_GENERAL_KEYS:
                # Will be logged after logging is configured; store for later.
                config.setdefault('_unknown_general', []).append(key)
                continue
            if value.strip():          # ignore blank values → use default
                config['general'][key] = value.strip()

    # ── [tabajara_con] ─────────────────────────────────────────────────────
    if parser_cfg.has_section('tabajara_con'):
        for key, value in parser_cfg.items('tabajara_con'):
            if value.strip():
                config['tabajara_con'][key] = value.strip()

    # ── [tabajara_dis] ─────────────────────────────────────────────────────
    if parser_cfg.has_section('tabajara_dis'):
        for key, value in parser_cfg.items('tabajara_dis'):
            if value.strip():
                config['tabajara_dis'][key] = value.strip()

    return config

def build_tabajara_args(params_dict, exclude_keys=None):
    """
    Converts a dict of tabajara parameters into a flat list of CLI arguments.

    Example: {'t': '0.5', 'p': '50', 'm': 'c'} -> ['-t', '0.5', '-p', '50', '-m', 'c']

    Keys in exclude_keys are skipped (used to avoid duplicating positional
    arguments that are already hard-coded in the caller, e.g. -i and -o).
    """
    exclude_keys = set(exclude_keys or [])
    args_list = []
    for key, value in params_dict.items():
        if key in exclude_keys:
            continue
        args_list.extend([f'-{key}', value])
    return args_list

def detect_cli_config_conflict(raw_argv, config_flag_used):
    """
    Checks whether the user passed -c/-config together with any parameter
    that is covered by the config file.  Aborts with a clear message if so.
    """
    if not config_flag_used:
        return

    conflicting = []
    for token in raw_argv:
        if token in CONFIG_COVERED_PARAMS:
            conflicting.append(token)

    if conflicting:
        print(
            f"\nError: Cannot use -c/-config together with the following "
            f"parameter(s): {', '.join(conflicting)}\n"
            f"Use EITHER the config file OR individual CLI flags, not both.\n"
        )
        sys.exit(1)

def _safe_path_name(name):
    """
    Sanitises a protein/marker name for use in file and directory names.

    Replaces characters that are unsafe in file-system paths (``/``,
    ``\\``, ``:``, ``*``, ``?``, ``"``, ``<``, ``>``, ``|``) and
    whitespace with underscores, then collapses consecutive underscores.

    Example
    -------
        _safe_path_name('Ser/Thr')         → 'Ser_Thr'
        _safe_path_name('RNA pol  II')     → 'RNA_pol_II'
    """
    sanitised = re.sub(r'[/\\:*?"<>|\s]+', '_', name)
    return sanitised.strip('_')

def _shorten_accession_filename(nuc_acc, limit=100):
    """
    Builds a base file name from a (possibly very long) 'Virus GENBANK
    accession' field, keeping it within the file-system limit.

    Segmented individuals may list many accessions in a single field (e.g.
    'partial: AY225133; partial: AY225134; ...'), which would produce a file
    name longer than the OS allows (ENAMETOOLONG / Errno 36). When the raw
    string exceeds *limit* characters, it is truncated to the first *limit*
    characters and suffixed with '_others_genbankID'. The result is
    deterministic (the same field always yields the same name).

    Example
    -------
        '...; partial: AY225148' (300+ chars)
            → '<first 100 chars>_others_genbankID'
    """
    name = str(nuc_acc)
    if len(name) > limit:
        name = name[:limit] + "_others_genbankID"
    return name

def search_entrez(nuc_acc):

    if isinstance(nuc_acc, float):
        return []
    # define ids
    if ":" or ";" in nuc_acc:
        ids = [
            identity for identity in nuc_acc.replace(";", " ").replace(":", " ").split()
            if len(identity) >= 6 and any(c.isalpha() for c in identity) and any(c.isdigit() for c in identity)
            ]
    else:
        ids = [nuc_acc]

    all_protein_ids = []

    # print(ids)
    for acc in ids:
        for attempt in range(3):
            try:
                with _ncbi_limiter:
                    handle = Entrez.esearch(db="nuccore", term=acc, usehistory='y')
                record = Entrez.read(handle)
                handle.close()

                webenv = record['WebEnv']
                query_key = record['QueryKey']

                with _ncbi_limiter:
                    search_elink = Entrez.elink(
                        dbfrom="nuccore",
                        db="protein",
                        query_key=query_key,
                        webenv=webenv,
                        linkname="nuccore_protein"
                    )

                result_elink = Entrez.read(search_elink)
                search_elink.close()

                if result_elink[0].get("LinkSetDb"):
                    for link in result_elink[0]["LinkSetDb"][0].get("Link", []):
                        all_protein_ids.append(link.get("Id", ""))

                break  # sucess → drop retry loop

            except (socket.timeout,
                    http.client.IncompleteRead,
                    ValueError,
                    RuntimeError,
                    http.client.RemoteDisconnected,
                    urllib.error.HTTPError,
                    urllib.error.URLError) as e:

                logging.warning(f"Error fetching data: {e}. attempt {attempt+1}/3")
                time.sleep(5)

        else:
            logging.error(f"Failed to retrieve data for {acc} after multiple attempts.")

    return all_protein_ids

def filtered_search(protein_ids, query_terms, negative_terms, min_len, max_len):

    for attempt in range(3):
        try:  
            filtered_ids = []

            if not protein_ids:
                return protein_ids 
            else:
                # Separate negative terms.
                separated_terms = []
                for item in negative_terms:
                    if pd.isna(item):
                        continue
                    separated_terms.extend([term.strip() for term in str(item).split(",")])

                len_search = f'"{int(float(min_len))}"[SLEN] : "{int(float(max_len))}"[SLEN]'

                query_terms = query_terms[0]
                if "," in query_terms:
                    positive_terms = query_terms.replace(",", " OR ")
                else:
                    positive_terms = query_terms


                search = f"({'[uid] OR '.join(protein_ids)}[uid]) AND ({positive_terms}) AND ({len_search}) NOT ({' OR '.join(separated_terms)})"
                # print(search)

                # filtered search in the protein database
                with _ncbi_limiter:
                    search_with_filter = Entrez.esearch(db="protein", term=search)
                filter_result = Entrez.read(search_with_filter)
                #print(resultado_filtro)
                search_with_filter.close()

                for id in filter_result.get("IdList", []):
                    filtered_ids.append(id)
                
                if not filtered_ids:
                    #print("No filtered ID found.)
                    return []
                
                with _ncbi_limiter:
                    efetch_handle = Entrez.efetch(db="protein", id=",".join(filtered_ids), rettype="gb", retmode="text")
                with efetch_handle as search_efetch:
                    response_data = search_efetch.read()

                seqio = SeqIO.parse(StringIO(response_data), "genbank")
                extracted_ids = [record.id for record in seqio]
                #print(f"Extracted IDs: {extracted_ids}")
                return extracted_ids[0]

            #return filtered_ids
        except (socket.timeout, http.client.IncompleteRead, ValueError, RuntimeError, http.client.RemoteDisconnected, urllib.error.HTTPError,urllib.error.URLError) as e:
            logging.warning(f"Error fetching data: {e}. attempts {attempt+1}/3...")
            time.sleep(5)
    logging.error("Failed to retrieve data after multiple attempts.")
    return []

def cds_prot(protein_ids,path,nuc_acc):   

    file_name = f"{_shorten_accession_filename(nuc_acc)}_prot.fasta"
    output_file = os.path.join(path, file_name)

    logging.info(f"Downloading all proteins coded by {nuc_acc} sequence…")
    for attempt in range(3):
        try:
            fasta_list = []
            if os.path.exists(output_file):
                #print(f"The file {output_file} already exists. Using the existing file.")
                return output_file
            elif protein_ids == []:
                return fasta_list
            else:
                
                with _ncbi_limiter:
                    efetch_h = Entrez.efetch(db="protein", id=",".join(protein_ids), rettype="fasta", retmode="text")
                with efetch_h as negative_result:
                    read_negative = negative_result.read()
                    negative_seqio = SeqIO.parse(StringIO(read_negative), "fasta")
                    
                    for seq_record in negative_seqio:
                        fasta_list.append(seq_record)

                    with open(output_file, "w") as output_write:
                        SeqIO.write(fasta_list, output_write, "fasta")
                #time.sleep(1) 
                #return fasta_records  # Returns the FASTA records
                return output_file

        except (socket.timeout, http.client.IncompleteRead, ValueError, RuntimeError, http.client.RemoteDisconnected, urllib.error.HTTPError, urllib.error.URLError) as e:
                logging.warning(f"Error fetching data: {e}. attempts {attempt+1}/3...")
                time.sleep(5)
    logging.error("Failed to retrieve data after multiple attempts.")
    return []

# Downloads the FASTA files of proteins from a family.
def refdb(name_protein, taxid, name_family, path, protein_name, min_len, max_len):

    
    # Formats the output file name.
    if "," in name_protein:
        main_name = name_protein.split(",")[0].strip()
    else:
        main_name = protein_name.strip()

    name = _safe_path_name(main_name)

    file_name = f"{name_family}_{name}.fasta"

    
    output_file = os.path.join(path, file_name)

    # Checks if the file already exists.
    if os.path.exists(output_file):
        # print(f"The file {output_file} already exists. Using the existing file.")
        return output_file

    logging.info(f'Building a BLAST database with reference protein sequences of {name_protein} of {name_family}')

    len_search = f'"{int(float(min_len))}"[SLEN] : "{int(float(max_len))}"[SLEN]'
    if "," in name_protein:
        positive_terms = name_protein.replace(",", " OR ")
    else:
        positive_terms = name_protein
    search = f"{positive_terms} AND txid{taxid}[Organism] AND {len_search}"
    print(search)

    # ── Helper: run esearch + batch efetch against a given NCBI db ────
    # Returns the list of SeqRecords found (empty list on failure or no
    # results). Uses the history server (WebEnv/QueryKey) so all hits are
    # fetched in a single efetch call.
    def _fetch_records_for_db(db):
        rec = None
        for attempt in range(3):
            try:
                with _ncbi_limiter:
                    handle = Entrez.esearch(db=db, term=search, retmax=200, usehistory='y')
                rec = Entrez.read(handle)
                handle.close()
                break
            except (socket.timeout,
                    http.client.IncompleteRead,
                    ValueError,
                    RuntimeError,
                    http.client.RemoteDisconnected,
                    urllib.error.HTTPError,
                    urllib.error.URLError) as e:
                logging.warning(f"Error fetching data (refdb esearch, db={db}) for {name_protein}: {e}. attempt {attempt+1}/3")
                time.sleep(5)

        if rec is None:
            logging.error(f"Failed to retrieve refdb esearch data (db={db}) for {name_protein} after multiple attempts.")
            return []

        if not rec.get('IdList') or int(rec.get('Count', 0)) == 0:
            return []

        webenv    = rec['WebEnv']
        query_key = rec['QueryKey']

        # Single batch efetch using the NCBI history server: reuse the
        # WebEnv/QueryKey from esearch to fetch all results in one request.
        records = []
        for attempt in range(3):
            try:
                with _ncbi_limiter:
                    efetch_handle = Entrez.efetch(
                        db=db,
                        query_key=query_key,
                        webenv=webenv,
                        rettype="fasta",
                        retmode="text",
                        retmax=200
                    )
                with efetch_handle as ef:
                    response_data = ef.read()

                # response may be bytes or str depending on Biopython version
                if isinstance(response_data, bytes):
                    response_data = response_data.decode('utf-8')

                records = list(SeqIO.parse(StringIO(response_data), "fasta"))
                break  # success
            except (socket.timeout,
                    http.client.IncompleteRead,
                    ValueError,
                    RuntimeError,
                    http.client.RemoteDisconnected,
                    urllib.error.HTTPError,
                    urllib.error.URLError) as e:
                logging.warning(f"Error fetching data (refdb batch efetch, db={db}): {e}. attempt {attempt+1}/3...")
                records = []
                time.sleep(5)

        return records

    # ── Primary attempt: 'ipg' database ───────────────────────────────
    fasta_records = _fetch_records_for_db("ipg")

    # ── Fallback: 2 or fewer sequences retrieved from 'ipg' → retry
    # against the 'protein' database. Only switch to the 'protein' result
    # if it yields strictly MORE sequences than 'ipg'; on a tie or fewer,
    # keep the original 'ipg' result. ─────────────────────────────────
    if len(fasta_records) <= 2:
        logging.info(
            f"Only {len(fasta_records)} sequence(s) from 'ipg' for "
            f"{name_protein} of {name_family}; retrying with db='protein'."
        )
        protein_records = _fetch_records_for_db("protein")
        if len(protein_records) > len(fasta_records):
            logging.info(
                f"Using 'protein' result ({len(protein_records)} seqs) over "
                f"'ipg' ({len(fasta_records)} seqs) for {name_protein}."
            )
            fasta_records = protein_records
        else:
            logging.info(
                f"'protein' result ({len(protein_records)} seqs) not better "
                f"than 'ipg' ({len(fasta_records)} seqs); keeping 'ipg'."
            )

    if not fasta_records:
        logging.warning(f"No FASTA records retrieved for {name_protein} of {name_family}.")
        return []  

    with open(output_file, "w") as output_handle:
        SeqIO.write(fasta_records, output_handle, "fasta")

    return output_file


# ──────────────────────────────────────────────────────────────────────────────
# Whole-genome download  (-gb yes)
# ──────────────────────────────────────────────────────────────────────────────

def download_genome_fasta(seq_number, genome_code, family, genome_data_dir):
    """
    Downloads the complete nucleotide genome (FASTA) of ONE VMR/MSL
    individual and saves it under genome_data/<family>/.

    All segments of a segmented genome are fetched in a single efetch call
    and written to one file. The file is named:

        VMR<seq_number:07d>_<acc1-acc2-...>.fasta

    where the accession part joins every accession found in the
    'Virus GENBANK accession' field (version suffixes removed) with '-'.

    Existing files are left untouched (allows resuming). Returns the output
    path on success, or None on skip/failure.
    """
    prefix = f"VMR{seq_number:07d}"

    fam = family if isinstance(family, str) and family.strip() else 'unclassified'
    fam_dir = os.path.join(genome_data_dir, _safe_path_name(fam))
    os.makedirs(fam_dir, exist_ok=True)

    # Validate accession field
    if (genome_code is None
            or isinstance(genome_code, float)
            or not str(genome_code).strip()
            or str(genome_code).strip().lower() == 'nan'):
        logging.warning(f"[genome] {prefix}: empty/invalid accession; skipped.")
        return None

    # Extract accession IDs (same parsing used by search_entrez; handles the
    # segmented ';'/':' syntax, e.g. 'L: MK861116; M: MK861117').
    raw = str(genome_code)
    acc_ids = [
        tok for tok in raw.replace(";", " ").replace(":", " ").split()
        if len(tok) >= 6 and any(c.isalpha() for c in tok) and any(c.isdigit() for c in tok)
    ]
    if not acc_ids:
        logging.warning(f"[genome] {prefix}: no valid accession in '{genome_code}'; skipped.")
        return None

    # Filename accession part: strip version suffix (.1) and join with '-'
    name_accs = _shorten_accession_filename("-".join(a.split('.')[0] for a in acc_ids))
    output_file = os.path.join(fam_dir, f"{prefix}_{name_accs}.fasta")

    if os.path.exists(output_file):
        logging.info(f"[genome] {os.path.basename(output_file)} already exists; skipped.")
        return output_file

    for attempt in range(3):
        try:
            with _ncbi_limiter:
                efetch_handle = Entrez.efetch(
                    db="nuccore",
                    id=",".join(acc_ids),
                    rettype="fasta",
                    retmode="text"
                )
            with efetch_handle as h:
                data = h.read()

            if isinstance(data, bytes):
                data = data.decode('utf-8')

            records = list(SeqIO.parse(StringIO(data), "fasta"))
            if not records:
                logging.warning(f"[genome] {prefix}: no sequence returned for {acc_ids}. attempt {attempt+1}/3")
                time.sleep(5)
                continue

            with open(output_file, "w") as out:
                SeqIO.write(records, out, "fasta")
            logging.info(f"[genome] Saved {os.path.basename(output_file)} ({len(records)} sequence(s)).")
            return output_file

        except (socket.timeout,
                http.client.IncompleteRead,
                ValueError,
                RuntimeError,
                http.client.RemoteDisconnected,
                urllib.error.HTTPError,
                urllib.error.URLError) as e:
            logging.warning(f"[genome] Error downloading {prefix} ({acc_ids}): {e}. attempt {attempt+1}/3")
            time.sleep(5)

    logging.error(f"[genome] Failed to download {prefix} ({acc_ids}) after 3 attempts.")
    return None


def download_all_genomes(tableX, final_dir, num_threads):
    """
    Iterates over every row (individual) of the VMR/MSL table and downloads
    its complete genome via download_genome_fasta.

    A sequential counter (1-based, in table order) is assigned to each row
    and rendered as VMR<counter:07d>. Family is resolved with the same
    fallback used elsewhere: Family → Subfamily → 'unclassified'.

    Runs in parallel when num_threads is set (reusing the -thr pool and the
    NCBI rate limiter), otherwise sequentially.
    """
    genome_data_dir = os.path.join(final_dir, "genome_data")
    os.makedirs(genome_data_dir, exist_ok=True)
    logging.info(f"[genome] Downloading viral genomes into {genome_data_dir} ...")

    tasks = []
    for i in range(len(tableX)):
        row         = tableX.iloc[i]
        genome_code = row.get("Virus GENBANK accession", "")
        family      = row.get("Family", "")
        subfamily   = row.get("Subfamily", "")

        if isinstance(family, str) and family.strip():
            fam = family
        elif isinstance(subfamily, str) and subfamily.strip():
            fam = subfamily
        else:
            fam = 'unclassified'

        tasks.append((i + 1, genome_code, fam))   # 1-based counter

    if num_threads is not None:
        with ThreadPoolExecutor(max_workers=num_threads) as executor:
            futures = [
                executor.submit(download_genome_fasta, n, gc, fam, genome_data_dir)
                for (n, gc, fam) in tasks
            ]
            for _ in as_completed(futures):
                pass
    else:
        for (n, gc, fam) in tasks:
            download_genome_fasta(n, gc, fam, genome_data_dir)

    logging.info(f"[genome] Genome download step finished ({len(tasks)} individual(s) processed).")


#Executes the commands of the blast_plus function.
def auxiliary(command):
    """Executes a command and checks if it was successful."""
    try:
        result = subprocess.run(command, check=True, text=True)
        logging.info(f"Command executed successfully: {' '.join(command)}")
        return result
    except subprocess.CalledProcessError as e:
        logging.error(f"Error: Failed to execute {' '.join(command)}")
        logging.error(e)
        return None

# Creates command lines that will be run in the shell.
def blast_plus(database_fasta,genome):

    result_file = f"Blast_result_{os.getpid()}_{int(time.time())}.txt"
    
    try:
        if genome == []:
             #print('The genome is a empty list')
            return None
        else:
            # print(f"Running BLAST: {genome} versus {database_fasta}")
            
            blastp_cmd = [
                "blastp",
                "-query", genome,
                "-db", database_fasta,
                "-out", result_file,
                "-evalue", "1e-20",
                "-num_alignments", "1",
                "-outfmt", ""'6 qseqid '""
            ]
            
            auxiliary(blastp_cmd)
        
        protein = None
        if os.path.exists(result_file):
            with open(result_file, "r") as read:
                lines = read.readlines()
                if lines:  # Checks if there are lines in the file.
                    protein = lines[0].strip()
                #     print(f"Protein found: {protein}")
                # else:
                #     print("No results found in the BLAST output file.")
        # else:
        #       print(f"The file {result_file} was not found.")
            
        return protein
            
    except Exception as e:
        logging.error(f"BLAST execution failed: {e}")
        return None
    finally:
        # This part will always be executed, ensuring the file is cleaned.
        if os.path.exists(result_file):
            try:
                os.remove(result_file)
                # print(f"Temporary file {result_file} was successfully removed..")
            except Exception as e:
                logging.error(f"Error removing temporary file {result_file}: {e}")

def make_blast_db(database_fasta):
    if os.path.exists(f'{database_fasta}.pdb'):
        return database_fasta
    else:
        makeblastdb_cmd = [
            "makeblastdb",
            "-in", database_fasta,
            "-dbtype", "prot",
            "-out", database_fasta
        ]
        auxiliary(makeblastdb_cmd)

def marker_fasta(fasta_file, keyword, path, genus, family):

    file_family = f"{family}.fasta"
    file_genus = f"{genus}.fasta"
    
    output_file = os.path.join(path, file_genus)
    allmarker_file = os.path.join(path, file_family)

    try:
        if keyword == []:
            return
        
        sequence = ""
        found = False
        
        with open(fasta_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
            for line in lines:
                if line.startswith('>'):
                    if found:
                        break
                    if keyword in line:
                        found = True  
                        sequence += line
                elif found:
                    sequence += line
        
        if found:
            if os.path.exists(output_file):
                with open(output_file, 'r', encoding='utf-8') as fasta_genus:
                    info_genus = fasta_genus.read()
                    if keyword not in info_genus:
                        with open(output_file, 'a', encoding='utf-8') as out:
                            out.write(sequence)
            else:
                with open(output_file, "w") as out:
                    out.write(sequence)
            
            line_sequence = sequence.strip().splitlines()
            

            prefixed_header = None
            prefixed_lines = []
            for line in line_sequence:
                if line.startswith('>'):
                    new_header = line.replace('>', f'>{genus}_', 1)
                    prefixed_header = new_header
                    prefixed_lines.append(new_header)
                else:
                    prefixed_lines.append(line)
            
            prefixed_sequence = "\n".join(prefixed_lines)

            if os.path.exists(allmarker_file):
                with open(allmarker_file, 'r', encoding='utf-8') as f:
                    information = f.read()
                    if prefixed_header is None or prefixed_header not in information:
                        with open(allmarker_file, 'a', encoding='utf-8') as out:
                            out.write("\n" + prefixed_sequence)
            else:
                with open(allmarker_file, "w") as out:
                    out.write(prefixed_sequence)

    except FileNotFoundError:
        logging.error(f'Error: The file {fasta_file} was not found')
    except Exception:
        return
    
def counter_prot(fasta_file):
    # Reads the file line by line.
    counter = 0
    if fasta_file == []:
        logging.warning(f"Downloaded {counter} protein sequences. ")
        return counter
    else:
        with open(fasta_file, 'r', encoding='utf-8') as read_file:
            for line in read_file:
                counter += line.count('>')
        logging.info(f"Downloaded {counter} protein sequences. ")
        return counter

def unique_dir(base_dir):
    if not os.path.exists(base_dir):
        return base_dir

    counter = 2
    new_dir = f"{base_dir}{counter}"
    while os.path.exists(new_dir):
        counter += 1
        new_dir = f"{base_dir}{counter}"
    return new_dir
    
def hyperlink_Protein_ID(protein_id):
    try:
        if protein_id == [] or protein_id is None:
            return 
        else:
            protein_id = str(protein_id).strip()
            hyperlink = f"https://www.ncbi.nlm.nih.gov/protein/{protein_id}"
            return hyperlink
    except Exception:
        return 
    
def hyperlink_code(genome_code):
    try:
        code = str(genome_code).strip()
        hyperlink = f"https://www.ncbi.nlm.nih.gov/nuccore/{code}"
        return hyperlink
    except Exception:
        return

def hyperlink_extrator(ws, col_name, ictv_id):

    code_ictv = str(ictv_id).strip()
    
    data = []  
    
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == col_name:
            letra = get_column_letter(col)
            hyperlink_form_default = re.compile(r'HYPERLINK\("([^"]+)"(?:,\s*"([^"]*)")?\)', re.IGNORECASE)
            
            for cell in ws[letra][1:]:
                if cell.value and code_ictv in str(cell.value):
                    link = None  

                    if cell.hyperlink:
                        link = cell.hyperlink.target
                    elif cell.data_type == 'f' and cell.value:
                        match = hyperlink_form_default.search(cell.value)
                        if match:
                            link = match.group(1)
                    if link:
                        data.append(link)
            break  
    
    if data:
        return data[0]
    else:
        return None


def hyperlink_code_seg(genome_code):

    if not genome_code or genome_code == [] or pd.isna(genome_code) or str(genome_code).strip().lower() == 'nan':
        return 

    genome_code_str = str(genome_code).strip()

    acc_ids = []

    for part in re.split(r';', genome_code_str):
        part = part.strip()
        if not part:
            continue


        if ':' in part:
            _, acc = part.split(':', 1)
            acc = acc.strip()
        else:
            acc = part.strip()

        acc = acc.split('.')[0].strip()

        if (len(acc) >= 4
                and any(c.isalpha() for c in acc)
                and any(c.isdigit() for c in acc)):
            acc_ids.append(acc)

    if not acc_ids:
        return None

    return f"https://www.ncbi.nlm.nih.gov/nuccore/{','.join(acc_ids)}"

def hyperlink_segment(seg_name, genome_code):
    """
    Builds a URL for a named segment of a segmented genome.

    Given the segment label (e.g. 'L') and the full genome_code field
    (e.g. 'L: MK861116; M: MK861117; S: MK861118' or 'MK861116; MK861117'),
    resolves the corresponding GenBank accession ID and returns the NCBI
    nuccore URL.

    Used only for segmented genomes (when genome_code contains ':' or ';').

    Parameters
    ----------
    seg_name    : str — the segment name returned by segment_named (e.g. 'L',
                  'Seg1')
    genome_code : str — the full 'Virus GENBANK accession' field

    Returns
    -------
    str | None  — the NCBI URL, or None if the ID cannot be resolved.
    """
    if not seg_name or not genome_code:
        return None

    genome_code_str = str(genome_code).strip()

    # ── Named format: 'L: MK861116; M: MK861117' ──────────────────────────
    if ':' in genome_code_str:
        for item in genome_code_str.split(';'):
            item = item.strip()
            if not item:
                continue
            try:
                name, acc = item.split(':', 1)
            except ValueError:
                continue
            if name.strip() == seg_name:
                acc_id = acc.strip().split('.')[0]  # remove version suffix
                return f"https://www.ncbi.nlm.nih.gov/nuccore/{acc_id}"

    # ── Unnamed format: positional (Seg1, Seg2, …) ────────────────────────
    # seg_name_named returns 'SegN' (1-based) when no ':' separator is found.
    seg_match = re.match(r'^Seg(\d+)$', str(seg_name))
    if seg_match:
        index = int(seg_match.group(1)) - 1   # convert to 0-based
        ids_list = [
            i.strip() for i in genome_code_str.split(';') if i.strip()
        ]
        if 0 <= index < len(ids_list):
            acc_id = ids_list[index].split('.')[0]
            return f"https://www.ncbi.nlm.nih.gov/nuccore/{acc_id}"

    return None


def run_mafft(fasta_family,output):

    try:
        mafft_cmd = [
            "mafft-linsi",
            "--maxiterate", "1000",
            "--thread", "20",
            "--reorder",
            "--quiet",
            fasta_family
        ]

        with open(output, "w") as out_f:
            result = subprocess.run(
                mafft_cmd,
                stdout=out_f,
                stderr=subprocess.PIPE,
                text=True,
                check=True
            )

        logging.info(f"Mafft alignment completed successfully")
        return result
    
    except Exception as e:
        logging.error(f"Error running MAFFT: {e}")
        return None


# ──────────────────────────────────────────────────────────────────────────────
# Redundancy padding  (make small alignments viable for HMM building)
# ──────────────────────────────────────────────────────────────────────────────

# Number of DUPLICATES to add as a function of how many real sequences a
# marker FASTA has. The idea is to DOUBLE the sequences (x2) repeatedly until
# the total exceeds 5, so MAFFT/tabajara can build informative profile HMMs:
#     1 seq -> x2 x2 x2 = 8   (add 7)
#     2 seq -> x2 x2    = 8   (add 6)
#     3 seq -> x2       = 6   (add 3)
#     >=4    -> unchanged
# The balanced cycling in _build_unique_duplicates reproduces the doubling
# (every original ends up with the same number of copies).
DUPLICATES_BY_COUNT = {1: 7, 2: 6, 3: 3}


def _num_duplicates(n):
    """Return how many duplicate records to add for *n* real sequences."""
    return DUPLICATES_BY_COUNT.get(n, 0)


def _read_fasta_blocks(path):
    """
    Reads a FASTA file and returns a list of records, each as a single string
    that starts with '>' and includes its sequence lines (newline-terminated).
    Headers and sequence content are preserved exactly.
    """
    blocks = []
    current = []
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.startswith('>'):
                if current:
                    blocks.append(''.join(current))
                    current = []
                current.append(line if line.endswith('\n') else line + '\n')
            elif current:
                current.append(line if line.endswith('\n') else line + '\n')
    if current:
        blocks.append(''.join(current))
    return blocks


def _seq_body(block):
    """Return the sequence lines (everything after the header line) of a block."""
    nl = block.find('\n')
    body = block[nl + 1:] if nl != -1 else ''
    if body and not body.endswith('\n'):
        body += '\n'
    return body


def _build_unique_duplicates(blocks, genus):
    """
    Builds the duplicate records to add for these *blocks*, cycling through
    their sequence bodies in a balanced way. The number of duplicates follows
    DUPLICATES_BY_COUNT (1->4, 2->3, 3->2, else 0). Each duplicate gets a
    UNIQUE, genus-prefixed header following the pattern:

        >{genus}_sequence, >{genus}_sequence2, >{genus}_sequence3, ...

    The genus prefix keeps tabajara's per-genus discrimination working, while
    the running suffix guarantees unique names (hmmbuild rejects duplicates).
    Returns a list of raw text blocks, or [] if no padding is needed.
    """
    n = len(blocks)
    d = _num_duplicates(n)
    if d == 0:
        return []
    dups = []
    for k in range(d):
        body   = _seq_body(blocks[k % n])
        suffix = "sequence" if k == 0 else f"sequence{k + 1}"
        dups.append(f">{genus}_{suffix}\n{body}")
    return dups


def pad_marker_fastas(dir_marker, redundancy_status):
    """
    Pads every marker's per-genus and per-family FASTA that has 1..3 real
    sequences by doubling them until the total exceeds 5 (see
    DUPLICATES_BY_COUNT: 1->8, 2->8, 3->6), so the downstream MAFFT alignment
    and tabajara HMM building remain viable when only a handful of real
    sequences were found.

    For each markers/<family>/<marker>/fasta/ folder:
      * every {genus}.fasta with 1..3 records is padded by duplicating
        its own sequences, each duplicate receiving a unique genus-prefixed
        header (>{genus}_sequence, >{genus}_sequence2, ...);
      * the SAME duplicate records are appended to {family}.fasta, so the
        family alignment reflects them and tabajara still resolves their genus
        from the header prefix;
      * as a safety net, if {family}.fasta itself has 1..3 records it is
        padded directly (prefix >{family}_sequence...).

    Populates *redundancy_status* (a dict) so the report can flag each group:
      (family, marker, genus) -> bool   # per-genus rows
      (family, marker, None)  -> bool   # the family-wide row
    """
    dir_marker = str(dir_marker)
    if not os.path.isdir(dir_marker):
        return

    def _append_blocks(path, blocks):
        with open(path, 'a', encoding='utf-8') as fh:
            for b in blocks:
                fh.write(b if b.startswith('\n') else '\n' + b)

    for family in sorted(os.listdir(dir_marker)):
        family_path = os.path.join(dir_marker, family)
        if not os.path.isdir(family_path):
            continue

        for marker in sorted(os.listdir(family_path)):
            marker_path = os.path.join(family_path, marker)
            if not os.path.isdir(marker_path):
                continue

            fasta_dir = os.path.join(marker_path, "fasta")
            if not os.path.isdir(fasta_dir):
                continue

            family_fasta = os.path.join(fasta_dir, f"{family}.fasta")
            family_had_dup = False

            # ── Per-genus FASTAs ──────────────────────────────────────────
            for fname in sorted(os.listdir(fasta_dir)):
                if not fname.endswith(".fasta") or fname == f"{family}.fasta":
                    continue
                genus = fname[:-len(".fasta")]
                genus_fasta = os.path.join(fasta_dir, fname)

                blocks = _read_fasta_blocks(genus_fasta)
                dups = _build_unique_duplicates(blocks, genus)

                if not dups:
                    redundancy_status[(family, marker, genus)] = False
                    continue

                # Same unique-header duplicates go to BOTH the genus and the
                # family FASTA (they already carry the genus prefix).
                _append_blocks(genus_fasta, dups)
                if os.path.exists(family_fasta):
                    _append_blocks(family_fasta, dups)

                redundancy_status[(family, marker, genus)] = True
                family_had_dup = True
                logging.info(
                    f"[redundancy] {family}/{marker}/{fname} padded from "
                    f"{len(blocks)} to {len(blocks) + len(dups)} sequence(s)."
                )

            # ── Safety net: family FASTA still below the minimum ──────────
            if os.path.exists(family_fasta):
                fam_blocks = _read_fasta_blocks(family_fasta)
                fam_dups = _build_unique_duplicates(fam_blocks, family)
                if fam_dups:
                    _append_blocks(family_fasta, fam_dups)
                    family_had_dup = True
                    logging.info(
                        f"[redundancy] {family}/{marker}/{family}.fasta padded from "
                        f"{len(fam_blocks)} to {len(fam_blocks) + len(fam_dups)} sequence(s)."
                    )

            redundancy_status[(family, marker, None)] = family_had_dup

def rename_hmm_files(valid_hmms_dir, fasta_family_path):
    """
    Renames all .hmm files in *valid_hmms_dir* by prepending the family name
    extracted from *fasta_family_path* (everything before the first '_').

    Example
    -------
    fasta_family_path : 'Peribunyaviridae_aligned.fasta'
      → prefix : 'Peribunyaviridae'

    Before : _6_663-700.hmm
    After  : Peribunyaviridae_6_663-700.hmm

    Already-prefixed files are skipped (idempotent).
    """
    valid_hmms_dir    = Path(valid_hmms_dir)
    fasta_family_path = Path(fasta_family_path)

    if not valid_hmms_dir.is_dir():
        logging.warning(f"rename_hmm_files: directory not found: {valid_hmms_dir}")
        return

    stem   = fasta_family_path.stem   # e.g. 'Peribunyaviridae_aligned'
    prefix = stem.split("_")[0]       # e.g. 'Peribunyaviridae'

    renamed = 0
    for hmm_file in sorted(valid_hmms_dir.glob("*.hmm")):
        old_name = hmm_file.name
        if old_name.startswith(prefix):   # already renamed — skip
            continue
        new_name = f"{prefix}{old_name}"
        hmm_file.rename(valid_hmms_dir / new_name)
        logging.info(f"  Renamed HMM: {old_name} → {new_name}")
        renamed += 1

    logging.info(f"rename_hmm_files: {renamed} file(s) renamed in {valid_hmms_dir}")

def write_tabajara_conf(params_dict, class_genus, conf_path):
    """
    Cria um arquivo de configuração para o Tabajara (.conf).

    O arquivo contém todos os parâmetros do dict mais o parâmetro 'c'
    (lista de gêneros). As chaves 'i' e 'o' são excluídas pois são
    passadas diretamente na linha de comando.

    Parâmetros
    ----------
    params_dict : dict  — parâmetros Tabajara (ex: TABAJARA_CON_DEFAULTS)
    class_genus : str   — string de gêneros separados por vírgula
    conf_path   : str   — caminho completo do arquivo a ser criado

    Retorna
    -------
    str — caminho do arquivo criado
    """
    with open(conf_path, 'w', encoding='utf-8') as f:
        f.write(f"c={class_genus}\n")

        for key, value in params_dict.items():
            if key == 'c':
                continue
            f.write(f"{key}={value}\n")

        logging.info(f"Tabajara config file created: {conf_path}")
        return conf_path


def run_tabajara_con(fasta_family, class_genus, output_dir, tabajara_params=None):
    """
    Runs tabajara.pl in conservative mode using a temporary .conf file.
    """
    output_con = output_dir / "tabajara_family"
    conf_path  = output_dir / "tabajara.conf"

    try:
        params = dict(TABAJARA_CON_DEFAULTS)
        if tabajara_params:
            params.update(tabajara_params)

        params['i'] = str(fasta_family)
        params['o'] = str(output_con)

        write_tabajara_conf(params, class_genus, conf_path)

        tabajara_cmd_con = [
            "tabajara.pl",
            "-conf", str(conf_path),
        ]

        logging.info(f"Running tabajara (con) with config: {conf_path}")

        result = subprocess.run(
            tabajara_cmd_con,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )

        logging.info("Tabajara HMMs profile (con) completed successfully")
        return result

    except Exception as e:
        logging.error(f"Error running Tabajara (con): {e}")
        return None

    finally:
            if conf_path.exists():
                try:
                    conf_path.unlink()
                    logging.info(f"Tabajara config file removed: {conf_path}")
                except subprocess.CalledProcessError as e:
                    logging.error(f"Tabajara returned code {e.returncode}")
                    logging.error(f"STDOUT:\n{e.stdout}")
                    logging.error(f"STDERR:\n{e.stderr}")
                    return None

            # ── Rename HMMs in tabajara_family only ───────────────────────────
            rename_hmm_files(output_con / "hmms" / "valid_HMMs", fasta_family)


def run_tabajara_dis(fasta_family, class_genus, output_dir, tabajara_params=None):
    """
    Runs tabajara.pl in discriminatory mode using a temporary .conf file.
    """
    output_dis = output_dir / "tabajara_genera"
    conf_path  = output_dir / "tabajara.conf"

    try:
        params = dict(TABAJARA_DIS_DEFAULTS)
        if tabajara_params:
            params.update(tabajara_params)

        params['i'] = str(fasta_family)
        params['o'] = str(output_dis)

        write_tabajara_conf(params, class_genus, conf_path)

        tabajara_cmd_dis = [
            "tabajara.pl",
            "-conf", str(conf_path),
        ]
        logging.info(f"Running tabajara (dis) with config: {conf_path}")

        result = subprocess.run(
            tabajara_cmd_dis,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )

        logging.info("Tabajara HMMs profile (dis) completed successfully")
        return result

    except Exception as e:
        logging.error(f"Error running Tabajara (dis): {e}")
        return None

    finally:
        if conf_path.exists():
            try:
                conf_path.unlink()
                logging.info(f"Tabajara config file removed: {conf_path}")
            except subprocess.CalledProcessError as e:
                logging.error(f"Tabajara returned code {e.returncode}")
                logging.error(f"STDOUT:\n{e.stdout}")
                logging.error(f"STDERR:\n{e.stderr}")
                return None
                
def create_family_protein_mapping(new_tableX):
    """
    Cria um mapeamento de (familia, termo_positivo) -> familia.
    Chave composta garante unicidade quando familias diferentes
    compartilham strings de termos identicos.
    """
    family_protein_map = {}
    for entry in new_tableX:
        family = entry.get('Family', 'Unassigned')
        positive_terms = entry.get('Positive_Terms', '')
        if positive_terms and not pd.isna(positive_terms):
            key = (str(family).strip(), str(positive_terms).strip())
            family_protein_map[key] = str(family).strip()
    return family_protein_map

def group_proteins_by_family(positive_dict, family_protein_map):
    """
    Agrupa os resultados por familia.
    """
    family_groups = defaultdict(lambda: {
        'proteins': {},
        'family_totals': {'annotation': 0, 'similarity': 0, 'undetected': 0}
    })

    for (family, term), counts in positive_dict.items():
        family_groups[family]['proteins'][term] = {
            'annotation': counts[0],
            'similarity': counts[1],
            'undetected': counts[2]
        }
        family_groups[family]['family_totals']['annotation']  += counts[0]
        family_groups[family]['family_totals']['similarity']  += counts[1]
        family_groups[family]['family_totals']['undetected']  += counts[2]

    return dict(family_groups)

def generate_family_grouped_report(genome_counter, positive_dict, new_tableX):
    """
    Gera o relatorio final agrupado por familia de proteina.
    """
    logging.info('Final report')
    logging.info(f'Total # of genome processed sequences: {genome_counter}')

    family_protein_map = create_family_protein_mapping(new_tableX)
    family_groups = group_proteins_by_family(positive_dict, family_protein_map)

    total_annotation = 0
    total_similarity = 0
    total_undetected = 0

    sorted_families = sorted(family_groups.keys(), key=lambda x: (x == 'Unassigned', x))

    for family in sorted_families:
        family_data   = family_groups[family]
        family_totals = family_data['family_totals']

        logging.info(f'========== Family: {family} ==========')
        logging.info(f'Family Summary - Total detected by annotation terms: {family_totals["annotation"]}')
        logging.info(f'Family Summary - Total detected by similarity search: {family_totals["similarity"]}')
        logging.info(f'Family Summary - Total undetected: {family_totals["undetected"]}')

        for protein_term, protein_counts in family_data['proteins'].items():
            logging.info(f'{protein_term.split(",")[0].strip()}')
            logging.info(f' Detected by annotation terms: {protein_counts["annotation"]}')
            logging.info(f' Detected by similarity search: {protein_counts["similarity"]}')
            logging.info(f' Undetected: {protein_counts["undetected"]}')

        logging.info('')

        total_annotation += family_totals['annotation']
        total_similarity += family_totals['similarity']
        total_undetected += family_totals['undetected']

    logging.info('========== Overall Summary ==========')
    logging.info(f'Total number of proteins detected by annotation terms: {total_annotation}')
    logging.info(f'Total number of proteins detected by similarity search: {total_similarity}')
    logging.info(f'Total number of undetected proteins: {total_undetected}')

def retrieve_genbank_accession(protein_id):

    for attempt in range(3):
        try:
            with _ncbi_limiter:
                handle = Entrez.efetch(
                    db="protein",
                    id=protein_id,
                    rettype="gb",
                    retmode="text"
                )
            data = handle.read()
            handle.close()

            record = SeqIO.read(StringIO(data), "genbank")

            db_source = record.annotations.get("db_source", "")
            if db_source:
                parts = db_source.split()
                for i, part in enumerate(parts):
                    if part.lower() == "accession" and i + 1 < len(parts):
                        accession = parts[i + 1].strip().rstrip(".")
                        return accession

            for feature in record.features:
                if feature.type == "CDS":
                    coded_by = feature.qualifiers.get("coded_by", [""])[0]
                    if coded_by:
                        acc = coded_by.split(":")[0].strip()
                        if acc.startswith("complement("):
                            acc = acc.replace("complement(", "")
                        if acc.startswith("join("):
                            acc = acc.replace("join(", "")
                        return acc

            return None

        except (socket.timeout,
                http.client.IncompleteRead,
                http.client.RemoteDisconnected,
                urllib.error.HTTPError,
                urllib.error.URLError,
                ValueError,
                RuntimeError) as e:
            logging.error(f"Error: {e}. Attempt {attempt + 1}/3...")
            time.sleep(5)

    logging.error(f"Failed to retrieve data for {protein_id} after 3 attempts.")
    return None

def segment_named(codeid, syntax):

    if codeid is None:
        return None

    codeid_base = codeid.split(".")[0]

    if ":" in syntax:
        for item in syntax.split(";"):
            item = item.strip()
            if not item:
                continue
            try:
                name, current_id = item.split(":", 1)
            except ValueError:
                continue
            if current_id.strip() == codeid_base:
                return name.strip()
        return None
    else:
        ids_list = [i.strip() for i in syntax.split(";") if i.strip()]
        try:
            index = ids_list.index(codeid_base)
            return f"Seg{index + 1}"
        except ValueError:
            return None

# ──────────────────────────────────────────────────────────────────────────────
# HMM collection, renaming and traceability report
# ──────────────────────────────────────────────────────────────────────────────

def _parse_hmm_metadata(hmm_path):
    """
    Parses an HMM file and extracts metadata fields.

    Returns a dict with keys:
        'LENG'          -> str or ''
        'NSEQ'          -> str or ''
        'CUTOFF_SCORE'  -> str or ''   (empty when the field is absent)
    """
    metadata = {'LENG': '', 'NSEQ': '', 'CUTOFF_SCORE': ''}

    leng_re  = re.compile(r'^LENG\s+(\S+)', re.MULTILINE)
    nseq_re  = re.compile(r'^NSEQ\s+(\S+)', re.MULTILINE)
    cutoff_re = re.compile(r'^CUTOFF SCORE\s+(\S+)', re.MULTILINE)

    with open(hmm_path, 'r', encoding='utf-8') as f:
        content = f.read()

    m = leng_re.search(content)
    if m:
        metadata['LENG'] = m.group(1)

    m = nseq_re.search(content)
    if m:
        metadata['NSEQ'] = m.group(1)

    m = cutoff_re.search(content)
    if m:
        metadata['CUTOFF_SCORE'] = m.group(1)

    return metadata


def fetch_taxid_by_name(taxon_name, cache=None):
    """
    Queries NCBI Taxonomy via Entrez to resolve a taxon name to its TaxID.

    Uses an in-memory cache dict so each unique name is queried only once.
    Returns the TaxID as a string, or '' on failure.
    """
    if cache is None:
        cache = {}

    if taxon_name in cache:
        return cache[taxon_name]

    for attempt in range(3):
        try:
            with _ncbi_limiter:
                handle = Entrez.esearch(db="taxonomy", term=taxon_name)
            record = Entrez.read(handle)
            handle.close()

            id_list = record.get("IdList", [])
            if id_list:
                txid = id_list[0]
                cache[taxon_name] = txid
                logging.info(f"TaxID for {taxon_name}: {txid}")
                return txid
            else:
                logging.warning(f"No TaxID found for {taxon_name}")
                cache[taxon_name] = ''
                return ''

        except (socket.timeout,
                http.client.IncompleteRead,
                ValueError,
                RuntimeError,
                http.client.RemoteDisconnected,
                urllib.error.HTTPError,
                urllib.error.URLError) as e:
            logging.warning(f"Error fetching TaxID for {taxon_name}: {e}. Attempt {attempt+1}/3")
            time.sleep(5)

    logging.error(f"Failed to retrieve TaxID for {taxon_name} after multiple attempts.")
    cache[taxon_name] = ''
    return ''


def collect_and_rename_hmms(markers_dir, output_base_dir):
    """
    Collects all .hmm files from valid_HMMs directories (both tabajara_family
    and tabajara_genera), copies them into a single flat directory
    (output_base_dir/hmms/) with globally numbered names (vHMM_1.hmm,
    vHMM_2.hmm, ...), and overwrites the NAME field inside each HMM file.

    Also extracts LENG, NSEQ, and CUTOFF SCORE from each file and fetches
    the TaxID for each family via NCBI Entrez (cached per family).

    Returns a list of dicts suitable for building a traceability report.

    Traversal order: families sorted alphabetically, then markers sorted
    alphabetically, then tabajara_family before tabajara_genera.  Within
    tabajara_genera, genus sub-folders are also sorted alphabetically.
    """

    hmm_output_dir = os.path.join(output_base_dir, "hmms")
    os.makedirs(hmm_output_dir, exist_ok=True)

    markers_path = Path(markers_dir)
    traceability = []
    counter = 0
    taxid_cache = {}

    # Regex to match the NAME line inside an HMM file
    name_re = re.compile(r'^(NAME\s+)(.+)$', re.MULTILINE)

    for family_path in sorted(markers_path.iterdir()):
        if not family_path.is_dir():
            continue
        family_name = family_path.name

        # Fetch TaxID once per family
        family_txid = fetch_taxid_by_name(family_name, cache=taxid_cache)

        for marker_path in sorted(family_path.iterdir()):
            if not marker_path.is_dir():
                continue
            marker_name_underscore= marker_path.name
            marker_name= marker_name_underscore.replace("_", " ")

            # ── tabajara_family ───────────────────────────────────────────
            family_valid = marker_path / "tabajara_family" / "hmms" / "valid_HMMs"
            if family_valid.is_dir():
                for hmm_file in sorted(family_valid.glob("*.hmm")):
                    counter += 1
                    new_name = f"vHMM_{counter}"
                    new_filename = f"{new_name}.hmm"
                    dest = os.path.join(hmm_output_dir, new_filename)

                    shutil.copy2(str(hmm_file), dest)

                    # Extract metadata before rewriting NAME
                    meta = _parse_hmm_metadata(dest)

                    _rewrite_hmm_name(dest, new_name, name_re)

                    traceability.append({
                        'vHMM_ID': new_name,
                        'Taxon': family_name,
                        'Genus': 'Family-wide models',
                        'Protein': marker_name,
                        'Model type': 'conservative',
                        'Length': meta['LENG'],
                        '# of sequences': meta['NSEQ'],
                        'Cutoff score': meta['CUTOFF_SCORE'],
                        'TxID': family_txid,
                        'Original_file': hmm_file.name,
                    })

            # ── tabajara_genera ───────────────────────────────────────────
            genera_hmms_base = marker_path / "tabajara_genera" / "hmms"
            if genera_hmms_base.is_dir():
                for genus_path in sorted(genera_hmms_base.iterdir()):
                    if not genus_path.is_dir():
                        continue
                    genus_name = genus_path.name

                    genus_valid = genus_path / "valid_HMMs"
                    if not genus_valid.is_dir():
                        continue

                    for hmm_file in sorted(genus_valid.glob("*.hmm")):
                        counter += 1
                        new_name = f"vHMM_{counter}"
                        new_filename = f"{new_name}.hmm"
                        dest = os.path.join(hmm_output_dir, new_filename)

                        shutil.copy2(str(hmm_file), dest)

                        meta = _parse_hmm_metadata(dest)

                        _rewrite_hmm_name(dest, new_name, name_re)

                        traceability.append({
                            'vHMM_ID': new_name,
                            'Taxon': family_name,
                            'Genus': genus_name,
                            'Protein': marker_name,
                            'Model type': 'discriminatory',
                            'Length': meta['LENG'],
                            '# of sequences': meta['NSEQ'],
                            'Cutoff score': meta['CUTOFF_SCORE'],
                            'TxID': family_txid,
                            'Original_file': hmm_file.name,
                        })

    logging.info(f"Collected and renamed {counter} HMM files into {hmm_output_dir}")
    return traceability


def _rewrite_hmm_name(hmm_path, new_name, name_re):
    """
    Reads an HMM file, replaces the NAME field value with *new_name*,
    and writes the file back in place.
    """
    with open(hmm_path, 'r', encoding='utf-8') as f:
        content = f.read()

    new_content = name_re.sub(rf'\g<1>{new_name}', content, count=1)

    with open(hmm_path, 'w', encoding='utf-8') as f:
        f.write(new_content)


def generate_hmm_traceability_report(traceability, output_base_dir):
    """
    Generates a CSV and XLSX traceability report.

    Columns:
        vHMM_ID, Taxon, Genus, Protein, Model type, Length,
        # of sequences, TxID, Original_file

    The 'Cutoff score' column is included ONLY when at least one HMM
    file contains a CUTOFF SCORE value; otherwise it is omitted entirely.
    """
    if not traceability:
        logging.info("No HMM files found; skipping traceability report.")
        return

    # Decide whether to include Cutoff score column
    has_cutoff = any(entry.get('Cutoff score', '') != '' for entry in traceability)

    columns = ['vHMM_ID', 'Taxon', 'Genus', 'Protein', 'Model type',
               'Length', '# of sequences']

    if has_cutoff:
        columns.append('Cutoff score')

    columns.extend(['TxID', 'Original_file'])

    report_csv = os.path.join(output_base_dir, "hmm_library.csv")

    with open(report_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=columns, delimiter=';',
                                extrasaction='ignore')
        writer.writeheader()
        writer.writerows(traceability)

    # CSV -> XLSX
    report_csv_path = Path(report_csv)
    df_trace = pd.read_csv(report_csv_path, delimiter=';')
    xlsx_trace = report_csv_path.with_suffix('.xlsx')
    df_trace.to_excel(xlsx_trace, index=False)

    logging.info(f"HMM traceability report generated: {report_csv} / {xlsx_trace.name}")


# ──────────────────────────────────────────────────────────────────────────────
# NCBI Rate Limiter  (token-bucket, thread-safe)
# ──────────────────────────────────────────────────────────────────────────────

class NCBIRateLimiter:
    """
    Cross-process, thread-safe token-bucket rate limiter for NCBI Entrez calls.

    With an API key NCBI allows up to 10 requests/second *for that key*,
    regardless of how many processes or threads are issuing requests with
    it.  This class enforces `max_rate` calls/second across:

        * all worker threads within this process, AND
        * all OTHER simultaneous executions of this program on the same
          machine that also use this limiter (multiple VMR.py runs started
          in parallel, e.g. via separate `nohup` / cron jobs).

    This is done with a shared lock file on disk (`fcntl.flock`, advisory,
    POSIX). Only one thread/process at a time may hold the file lock; while
    holding it, the caller checks the timestamp of the last issued request
    (persisted in the file itself) and sleeps the remaining gap before
    updating the timestamp and releasing the lock. Because the wait happens
    *while the lock is held*, every consumer — in this process or any other
    — is naturally serialized to the configured rate.

    Usage
    -----
        limiter = NCBIRateLimiter(max_rate=10)   # 10 req/s (NCBI ceiling)
        with limiter:                             # blocks until a token is free
            handle = Entrez.esearch(...)

    The limiter is a no-op singleton when parallel mode is disabled
    (``NCBIRateLimiter(max_rate=0)``): the context manager returns
    immediately without acquiring any lock.
    """

    # Shared across ALL processes/executions on this machine. Override via
    # the VMR_NCBI_LOCKFILE environment variable if you need an isolated
    # lock (e.g. for testing) or a path on a faster/local filesystem.
    LOCK_FILE = os.environ.get("VMR_NCBI_LOCKFILE", "/tmp/.vmr_ncbi_ratelimit.lock")

    def __init__(self, max_rate: float):
        """
        Parameters
        ----------
        max_rate : float
            Maximum allowed requests per second across ALL threads AND all
            other concurrently running instances of this program that share
            the same lock file.
            Pass 0 (or any non-positive value) to create a disabled limiter
            that never blocks.
        """
        self._disabled  = (max_rate <= 0)
        if self._disabled:
            return

        self._max_rate   = float(max_rate)
        self._min_gap    = 1.0 / self._max_rate   # minimum seconds between tokens

        # Make sure the shared lock file exists before any thread/process
        # tries to open it with 'r+' (which requires the file to exist).
        try:
            if not os.path.exists(self.LOCK_FILE):
                with open(self.LOCK_FILE, 'a'):
                    pass
        except OSError as e:
            logging.warning(
                f"NCBIRateLimiter: could not create lock file {self.LOCK_FILE} "
                f"({e}). Falling back to in-process-only rate limiting."
            )

    # ── context manager interface ──────────────────────────────────────────

    def __enter__(self):
        if self._disabled:
            return self

        try:
            with open(self.LOCK_FILE, 'r+') as f:
                fcntl.flock(f, fcntl.LOCK_EX)
                try:
                    content = f.read().strip()
                    try:
                        last_call = float(content) if content else 0.0
                    except ValueError:
                        last_call = 0.0

                    now     = time.time()
                    elapsed = now - last_call
                    wait    = self._min_gap - elapsed
                    if wait > 0:
                        time.sleep(wait)

                    now = time.time()
                    f.seek(0)
                    f.truncate()
                    f.write(repr(now))
                    f.flush()
                    os.fsync(f.fileno())
                finally:
                    fcntl.flock(f, fcntl.LOCK_UN)
        except OSError as e:
            # Lock file became unavailable mid-run (deleted, permissions,
            # filesystem issue, ...). Degrade gracefully instead of crashing
            # the whole pipeline over a rate-limiting bookkeeping failure.
            logging.warning(
                f"NCBIRateLimiter: lock file unavailable ({e}). "
                f"Proceeding without cross-process throttling for this call."
            )

        return self

    def __exit__(self, *_):
        pass   # nothing to release


# Module-level singleton; replaced in __main__ after CLI parsing.
# NOTE: sem type annotation — Python proibe `global x` dentro de funcao
# quando x foi declarado com anotacao no escopo de modulo.
_ncbi_limiter = NCBIRateLimiter(max_rate=0)


def _rate_limited_entrez_call(fn, *args, **kwargs):
    """
    Wraps any callable that makes a single Entrez network call with the
    global rate limiter.  Use this inside functions that perform their own
    retry loops so the limiter is consulted on every individual attempt.

    Example
    -------
        handle = _rate_limited_entrez_call(Entrez.esearch, db="nuccore", term=acc)
    """
    with _ncbi_limiter:
        return fn(*args, **kwargs)


# ──────────────────────────────────────────────────────────────────────────────
# Entrez results cache  (avoids redundant NCBI calls for the same genome)
# ──────────────────────────────────────────────────────────────────────────────

class EntrezCache:
    """
    Thread-safe cache for NCBI Entrez results, keyed by genome_code.

    When multiple protein markers are processed for the same genome, the
    calls to ``search_entrez`` (esearch + elink) and ``cds_prot`` (efetch)
    would be repeated identically — same genome_code always yields the
    same set of protein IDs and the same CDS FASTA.  This cache ensures
    each genome_code is queried only once.

    Per-key locking guarantees that if two threads request the same
    genome_code simultaneously, only one performs the NCBI call while
    the other blocks and then reuses the result.

    Usage
    -----
        result = _entrez_cache.get_or_compute(
            genome_code, 'search',
            lambda: search_entrez(genome_code)
        )
    """

    def __init__(self):
        self._global_lock = threading.Lock()
        self._key_locks   = {}            # genome_code -> Lock
        self._data        = {}            # genome_code -> {field: value}

    def _get_key_lock(self, genome_code):
        with self._global_lock:
            if genome_code not in self._key_locks:
                self._key_locks[genome_code] = threading.Lock()
            return self._key_locks[genome_code]

    def get_or_compute(self, genome_code, field, compute_fn):
        """
        Returns the cached value for *(genome_code, field)* if it exists,
        otherwise calls *compute_fn()*, caches and returns the result.

        Parameters
        ----------
        genome_code : str   — cache key (the GENBANK accession string)
        field       : str   — sub-key (e.g. 'search', 'cds', 'nprot')
        compute_fn  : callable — zero-arg function that produces the value
        """
        key_lock = self._get_key_lock(genome_code)
        with key_lock:
            # Check cache (only this thread can hold key_lock for this key)
            with self._global_lock:
                entry = self._data.get(genome_code)
                if entry is not None and field in entry:
                    logging.info(
                        f"[cache] Reusing '{field}' for {genome_code}"
                    )
                    return entry[field]

            # Not cached — compute (may take seconds: NCBI call)
            result = compute_fn()

            # Store result
            with self._global_lock:
                if genome_code not in self._data:
                    self._data[genome_code] = {}
                self._data[genome_code][field] = result

            return result


# Module-level singleton; initialised in __main__ before the main loop.
_entrez_cache = None


# ──────────────────────────────────────────────────────────────────────────────
# Parallel worker  (one task = one (tableX row, tableY row) pair)
# ──────────────────────────────────────────────────────────────────────────────

def _process_single_task(task):
    """
    Executes all network-bound work for one (genome × protein) pair.

    This function is designed to run inside a ThreadPoolExecutor worker.
    It is a self-contained unit: it receives all data it needs via *task*,
    performs the Entrez queries (rate-limited), the refdb download,
    cds_prot download, BLAST search and marker extraction, and returns a
    result dict suitable for appending to new_tableX.

    On unrecoverable failure the function logs the error and returns None
    so the caller can record a partial result and continue.

    Parameters
    ----------
    task : dict with keys
        i, j               — row indices into tableX / tableY
        line               — dict copy of tableX.iloc[i]
        tableY_row         — dict of the matching tableY row
        ws                 — openpyxl worksheet (read-only in threads, safe)
        final_dir          — Path to the output directory
        dir_marker         — str path to markers root dir

    Returns
    -------
    dict | None
    """
    try:
        # ── Unpack task ───────────────────────────────────────────────────
        line              = task['line']
        ty                = task['tableY_row']
        ws                = task['ws']
        final_dir         = task['final_dir']

        genome_code       = line.get("Virus GENBANK accession", "")
        genus             = line.get("Genus", "")
        family            = line.get("Family", "")
        subfamily         = line.get("Subfamily", "")
        ictv_id           = line.get("ICTV_ID", "")

        if isinstance(family, str):
            pass
        elif isinstance(family, float) and isinstance(subfamily, str):
            family = subfamily
        elif isinstance(family, float) and isinstance(subfamily, float):
            family = 'unclassified'

        protein_name          = ty['positive_terms']
        definitive_prot_name  = protein_name
        negative_terms_val    = ty['negative_terms']
        min_len_val           = ty['min_len']
        max_len_val           = ty['max_len']
        txid_val              = ty['txid']

        if "," in protein_name:
            main_protein_name = protein_name.split(",")[0].strip()
        else:
            main_protein_name = protein_name.strip()

        underscore_protein_name = _safe_path_name(main_protein_name)

        logging.info(
            f"[parallel] Processing GenBank ID {genome_code} "
            f"with {definitive_prot_name} as annotation term"
        )

        # ── Build sub-directories (thread-safe: exist_ok=True) ───────────
        dir_genome       = os.path.join(final_dir, "cds_virus_fasta")
        dir_genome_final = os.path.join(dir_genome, family, genus)
        dir_refdb_info   = os.path.join(final_dir, "refdb")
        dir_refdb_final  = os.path.join(dir_refdb_info, family, underscore_protein_name)
        dir_marker       = os.path.join(final_dir, "markers")
        dir_marker_final = os.path.join(dir_marker, family, underscore_protein_name, "fasta")

        os.makedirs(dir_genome_final,  exist_ok=True)
        os.makedirs(dir_refdb_final,   exist_ok=True)
        os.makedirs(dir_marker_final,  exist_ok=True)

        # ── Hyperlinks (read-only worksheet access) ───────────────────────
        ictv_id_url = hyperlink_extrator(ws, "ICTV_ID", ictv_id)

        genome_code_valido = (
            genome_code
            and not pd.isna(genome_code)
            and str(genome_code).strip().lower() != 'nan'
            and str(genome_code).strip() != ''
        )

        if not genome_code_valido:
            genome_code_url = None
        else:
            is_segmented = (
                ":" in str(genome_code) or ";" in str(genome_code)
            )
            if is_segmented:
                genome_code_url = hyperlink_code_seg(genome_code)
            else:
                genome_code_url = hyperlink_code(genome_code)

        # ── Cached / rate-limited Entrez calls ────────────────────────────
        # search_entrez and cds_prot depend only on genome_code; when
        # multiple markers are processed for the same genome the results
        # are reused via _entrez_cache, avoiding redundant NCBI calls.
        Gi = _entrez_cache.get_or_compute(
            genome_code, 'search',
            lambda: search_entrez(genome_code)
        )
        protein = filtered_search(
            Gi, [protein_name], [negative_terms_val], min_len_val, max_len_val
        )
        protein_url = hyperlink_Protein_ID(protein)

        seg     = None
        seg_url = None

        if not (pd.isna(genome_code) or not genome_code) and is_segmented:
            if protein and protein != []:
                retrive_gen = retrieve_genbank_accession(protein)
                seg     = segment_named(retrive_gen, genome_code)
                seg_url = hyperlink_segment(seg, genome_code)

        fasta_file = _entrez_cache.get_or_compute(
            genome_code, 'cds',
            lambda: cds_prot(Gi, dir_genome_final, genome_code)
        )
        number_prot = _entrez_cache.get_or_compute(
            genome_code, 'nprot',
            lambda: counter_prot(fasta_file)
        )
        marker_fasta(fasta_file, protein, dir_marker_final, genus, family)

        # ── Annotation counters (returned; aggregated by caller) ──────────
        annotation_hit  = 0
        similarity_hit  = 0
        undetected      = 0

        if protein == []:
            logging.info(
                f'[parallel] No protein annotated as {definitive_prot_name} '
                f'found on GenBank ID {genome_code}.'
            )
        else:
            annotation_hit = 1
            logging.info(
                f'[parallel] Protein annotated as {definitive_prot_name} '
                f'found: {protein}'
            )

        # ── BLAST fallback ────────────────────────────────────────────────
        if protein == []:
            logging.info(
                f'[parallel] Running BLAST on {number_prot} proteins '
                f'vs reference DB for {definitive_prot_name}'
            )
            database = refdb(
                protein_name, txid_val, family,
                dir_refdb_final, protein_name, min_len_val, max_len_val
            )
            protein     = blast_plus(database, fasta_file)
            protein_url = hyperlink_Protein_ID(protein)

            seg     = None
            seg_url = None

            if not (pd.isna(genome_code) or not genome_code) and is_segmented:
                if protein and protein != []:
                    retrive_gen = retrieve_genbank_accession(protein)
                    seg     = segment_named(retrive_gen, genome_code)
                    seg_url = hyperlink_segment(seg, genome_code)

            marker_fasta(fasta_file, protein, dir_marker_final, genus, family)

            if protein is None:
                undetected = 1
                logging.info(f'[parallel] No hit found for {definitive_prot_name}.')
            else:
                similarity_hit = 1
                logging.info(
                    f'[parallel] Positive protein for {definitive_prot_name}: {protein}.'
                )

        # ── Assemble result row ───────────────────────────────────────────
        result_line = dict(line)
        result_line['min_length']        = min_len_val
        result_line['max_length']        = max_len_val
        result_line['Negative_Terms']    = negative_terms_val
        result_line['Positive_Terms']    = protein_name
        result_line['Protein_ID']        = protein if protein else ""
        result_line['Segment']           = seg if seg else ""
        result_line['tax_id']            = txid_val
        result_line['ICVT_ID link']      = ictv_id_url
        result_line['Accession link(s) per segment(s)'] = genome_code_url
        result_line['Protein ID link']   = protein_url if protein_url else ""
        result_line['Segment link']      = seg_url if seg_url else ""

        return {
            'row':            result_line,
            'family':         str(family).strip(),
            'positive_terms': str(protein_name).strip(),
            'annotation':     annotation_hit,
            'similarity':     similarity_hit,
            'undetected':     undetected,
        }

    except Exception as exc:
        logging.error(
            f"[parallel] Unrecoverable error for genome_code="
            f"{task.get('line', {}).get('Virus GENBANK accession', '?')} "
            f"/ protein={task.get('tableY_row', {}).get('positive_terms', '?')}: "
            f"{exc}",
            exc_info=True,
        )
        return None   # caller will skip and continue


def run_parallel_pipeline(
    tableX, tableY,
    positive_terms_s, negative_terms_s,
    min_len_s, max_len_s, txid_s, name_fam_s,
    ws, final_dir, positive_dict,
    num_workers,
):
    """
    Parallel entry point that replaces the sequential double for-loop over
    tableX × tableY.

    Builds one task per matching (i, j) pair, submits them to a
    ThreadPoolExecutor, collects results in original order, and returns
    (new_tableX, family_markers_list).

    Thread-safety notes
    -------------------
    * os.makedirs(..., exist_ok=True)  — safe for concurrent calls.
    * openpyxl worksheet (ws) — used read-only; safe.
    * positive_dict — updated by the *caller* after all futures resolve,
      so no concurrent writes occur.
    * logging — Python's logging module is thread-safe.
    * File writes (cds_prot, marker_fasta) — each thread writes to a
      unique path derived from genome_code + genus, so no collision.
    """

    # ── Build task list (preserves original order) ────────────────────────
    tasks = []
    for i in range(len(tableX)):
        for j in range(len(tableY)):
            if tableX.loc[i, 'Family'] != tableY.loc[j, 'Name']:
                continue
            tasks.append({
                'i':    i,
                'j':    j,
                'line': dict(tableX.iloc[i]),
                'tableY_row': {
                    'positive_terms': positive_terms_s[j],
                    'negative_terms': negative_terms_s[j],
                    'min_len':        min_len_s[j],
                    'max_len':        max_len_s[j],
                    'txid':           txid_s[j],
                    'name_fam':       name_fam_s[j],
                },
                'ws':        ws,
                'final_dir': final_dir,
            })

    logging.info(
        f"[parallel] Submitting {len(tasks)} tasks to "
        f"{num_workers} worker threads."
    )

    new_tableX      = []
    family_markers  = []

    # ── Submit and collect in order ───────────────────────────────────────
    # We use a list of futures indexed by task position so that the output
    # table preserves the same row order as the sequential version.
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        future_to_task = {
            executor.submit(_process_single_task, t): t
            for t in tasks
        }

        completed = 0
        for future in as_completed(future_to_task):
            completed += 1
            t = future_to_task[future]
            genome_label = t['line'].get('Virus GENBANK accession', '?')
            prot_label   = t['tableY_row']['positive_terms']

            try:
                result = future.result()
            except Exception as exc:
                logging.error(
                    f"[parallel] Future raised for genome={genome_label} "
                    f"protein={prot_label}: {exc}",
                    exc_info=True,
                )
                result = None

            if result is None:
                # Partial failure: log and skip this row
                logging.warning(
                    f"[parallel] Skipping row genome={genome_label} "
                    f"protein={prot_label} due to unrecoverable error."
                )
                continue

            # ── Aggregate counters ────────────────────────────────────────
            dict_key = (result['family'], result['positive_terms'])
            if dict_key in positive_dict:
                positive_dict[dict_key][0] += result['annotation']
                positive_dict[dict_key][1] += result['similarity']
                positive_dict[dict_key][2] += result['undetected']

            new_tableX.append(result['row'])
            family_markers.append(result['family'])

            logging.info(
                f"[parallel] Completed {completed}/{len(tasks)} — "
                f"genome={genome_label}"
            )

    # Sort results back to original tableX order (as_completed is unordered)
    # We use the original (i, j) indices preserved in the task dict to
    # reconstruct a stable sort key.
    task_order = {
        (t['i'], t['j']): idx for idx, t in enumerate(tasks)
    }

    def _sort_key(row):
        # Match row back to task by genome_code + Positive_Terms
        gc = row.get('Virus GENBANK accession', '')
        pt = row.get('Positive_Terms', '')
        for (i, j), idx in task_order.items():
            t = tasks[idx]
            if (t['line'].get('Virus GENBANK accession', '') == gc
                    and t['tableY_row']['positive_terms'] == pt):
                return idx
        return 999999

    new_tableX.sort(key=_sort_key)

    logging.info(
        f"[parallel] Pipeline complete. "
        f"{len(new_tableX)}/{len(tasks)} rows collected."
    )

    return new_tableX, family_markers


# ──────────────────────────────────────────────────────────────────────────────
# Argument parsing
# ──────────────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser(add_help=False, formatter_class=RawTextHelpFormatter)
parser.add_argument('-i')
parser.add_argument("-o", "-output", default='output_dir')
parser.add_argument('-h', '-help', action='store_true')
parser.add_argument('-v', '-version', action='store_true')
parser.add_argument('-t')
parser.add_argument('-s', type=int, default=1)
parser.add_argument('-ts', type=int, default=1)
parser.add_argument('-c', '-config', dest='config', default=None,
                    metavar='FILE',
                    help='Path to a .ini configuration file (optional)')
parser.add_argument('--generate-config', action='store_true',
                    help='Generate a template config file (VMR_config_template.ini) and exit')
parser.add_argument('-gb', dest='gb', default=None, metavar='yes/no',
                    help="When 'yes', download every genome (FASTA) from the "
                         "VMR/MSL table into 'genome_data/'. Default: no.")
parser.add_argument('-thread', '-thr', dest='threads', type=int, default=None,
                    metavar='N',
                    help='Number of parallel worker threads for NCBI calls. '
                         'Omit to run sequentially.')
args = parser.parse_args()


family_markers = []
genus_name = []

if __name__ == '__main__':

    # ── --generate-config: generate template and exit immediately ─────────
    if args.generate_config:
        generate_config_template()
        sys.exit(0)

    if not len(sys.argv) > 1:
        print(help_text)

    elif args.h:
        print(help_text)

    elif args.v:
        print(f"""
VMR Program version {version} - 30 jan 2026
(c) 2024. Rafael Santos da Silva & Arthur Gruber
""")

    else:

        # ── Conflict detection ────────────────────────────────────────────
        detect_cli_config_conflict(sys.argv[1:], config_flag_used=args.config is not None)

        # ── Load config (if supplied) or fall back to CLI args ────────────
        cfg_general     = {}
        cfg_tab_con     = None   # None → run_tabajara_con uses its own defaults
        cfg_tab_dis     = None   # None → run_tabajara_dis uses its own defaults

        if args.config:
            raw_cfg = load_config(args.config)

            cfg_general    = raw_cfg.get('general', {})
            cfg_tab_con    = raw_cfg.get('tabajara_con')
            cfg_tab_dis    = raw_cfg.get('tabajara_dis')

            # Resolve final parameter values from config
            input_file  = cfg_general.get('input') or None
            output_dir  = cfg_general.get('output', 'output_dir')
            sheet_num   = int(cfg_general.get('sheet', 1))
            term_file   = cfg_general.get('terms') or None
            term_sheet  = int(cfg_general.get('terms_sheet', 1))
            email_val   = cfg_general.get('email', Entrez.email)
            api_key_val = cfg_general.get('api_key', Entrez.api_key)
            gb_val      = cfg_general.get('gb', args.gb)

            if not input_file:
                print("Error: 'input' is required. Set it in the config file or use -i.")
                sys.exit(1)
            if not term_file:
                print("Error: 'terms' is required. Set it in the config file or use -t.")
                sys.exit(1)

        else:
            # Traditional CLI path
            input_file  = args.i
            output_dir  = args.o
            sheet_num   = args.s
            term_file   = args.t
            term_sheet  = args.ts
            email_val   = Entrez.email
            api_key_val = Entrez.api_key
            gb_val      = args.gb

            if not input_file:
                print("Error: input file is required. Use -i <file> or -c <config.ini>.")
                sys.exit(1)
            if not term_file:
                print("Error: terms file is required. Use -t <file> or -c <config.ini>.")
                sys.exit(1)

        # Whether to download all genomes (-gb yes / gb = yes in config)
        download_genomes = str(gb_val).strip().lower() == 'yes' if gb_val else False

        # Apply Entrez credentials (may have been overridden by config)
        Entrez.email   = email_val
        Entrez.api_key = api_key_val

        # ── Output directory ──────────────────────────────────────────────
        final_dir = unique_dir(output_dir)
        os.makedirs(final_dir)

        log_file = os.path.join(final_dir, "VMR.log")

        logging.basicConfig(
            filename=log_file,
            filemode='a',
            format='%(asctime)s - %(levelname)s - %(message)s',
            level=logging.INFO
        )

        # Emit deferred warnings for unknown [general] keys
        for unknown_key in raw_cfg.get('_unknown_general', []) if args.config else []:
            logging.warning(
                f"Config file: unknown key '{unknown_key}' in [general] section -- ignored."
            )

        # ── Parallel mode setup ───────────────────────────────────────────
        num_threads = args.threads   # None = sequential

        if num_threads is not None:
            if num_threads < 1:
                print("Error: -thread/-thr must be a positive integer (>= 1).")
                sys.exit(1)
            if num_threads > 10:
                logging.warning(
                    f"-thread value {num_threads} exceeds the NCBI API limit of "
                    f"10 req/s. Consider using -thread 10 or lower to avoid "
                    f"HTTP 429 errors."
                )
            if not Entrez.api_key:
                logging.warning(
                    "Parallel mode (-thread) is active but no NCBI API key is set. "
                    "Without an API key the NCBI limit is 3 req/s. "
                    "Set api_key in the config file or hardcode Entrez.api_key."
                )
            # Rate: with an API key NCBI's hard ceiling is 10 req/s.  The
            # token-bucket limiter below enforces that ceiling exactly, so
            # we no longer need to shave off extra headroom by capping at 8;
            # we just make sure not to exceed 10 req/s regardless of how
            # many threads are requested.
            max_rate = min(num_threads, 10)   # cap at the NCBI ceiling (10 req/s)
            _ncbi_limiter = NCBIRateLimiter(max_rate=max_rate)
            logging.info(
                f"Parallel mode enabled: {num_threads} worker thread(s), "
                f"rate limiter set to {max_rate} req/s."
            )
        else:
            logging.info("Sequential mode (use -thread N to enable parallelism).")

        start_time = time.perf_counter()
        logging.info("Starting execution...")
        if args.config:
            logging.info(f"Configuration loaded from: {args.config}")

        # ── VMR/MSL table → CSV ───────────────────────────────────────────
        sheet = sheet_num - 1
        wb = load_workbook(input_file)
        ws = wb.worksheets[sheet]

        df_xl = pd.read_excel(input_file, sheet_name=sheet)
        file = Path(input_file)
        new_file = file.with_suffix(".csv")
        csv_file = os.path.join(final_dir, new_file.name)
        df_xl.to_csv(csv_file, sep=";", index=False)
        logging.info(f"{file.name} to {new_file.name} conversion")

        # ── Terms table → CSV ─────────────────────────────────────────────
        term_sheet_idx = term_sheet - 1
        term_df_xl = pd.read_excel(term_file, sheet_name=term_sheet_idx)
        tfile = Path(term_file)
        term_new_file = tfile.with_suffix(".csv")
        term_csv_file = os.path.join(final_dir, term_new_file.name)
        term_df_xl.to_csv(term_csv_file, sep=";", index=False)
        logging.info(f"{tfile.name} to {term_new_file.name} conversion")

        tableX = pd.DataFrame(pd.read_csv(csv_file, delimiter=';'))
        tableY = pd.read_csv(term_csv_file, delimiter=';')
        genome_conter = tableX['Virus GENBANK accession'].nunique()

        txid = tableY['tax_id'] 
        positive_terms = tableY['Positive_terms']
        negative_terms = tableY['Negative_terms']
        min_len = tableY['min_length']
        max_len = tableY['max_length']
        parent = tableY['Parent']
        name_fam = tableY["Name"]

        positive_dict = {}
        for y_init in range(len(tableY)):
            key = (str(name_fam[y_init]).strip(), str(positive_terms[y_init]).strip())
            positive_dict[key] = [0, 0, 0]

        # ── Entrez results cache (avoids duplicate NCBI calls) ────────────
        _entrez_cache = EntrezCache()
        logging.info("Entrez results cache initialised.")

        # ── Optional whole-genome download step (-gb yes) ─────────────────
        # Runs before the marker pipeline; the normal pipeline continues
        # afterwards regardless.
        if download_genomes:
            download_all_genomes(tableX, final_dir, num_threads)

        def _build_refdb_for_term(y):
            """
            Builds the reference protein database (and its BLAST index) for
            a single row of the terms table.
            """
            name_family = name_fam[y]
            protein_name_refdb = positive_terms[y]

            if pd.isna(protein_name_refdb) or not str(protein_name_refdb).strip():
                logging.warning(
                    f"Row {y} in terms table has an empty 'Positive_terms' "
                    f"value (Name='{name_family}'). Skipping refdb build for this row."
                )
                return

            protein_name_refdb = str(protein_name_refdb)

            if "," in protein_name_refdb:
                main_protein_name_refdb = protein_name_refdb.split(",")[0].strip()
            else:
                main_protein_name_refdb = protein_name_refdb.strip()

            underscore_protein_name_refdb = _safe_path_name(main_protein_name_refdb)

            dir_refdb = os.path.join(final_dir, "refdb")
            dir_refdb_final = os.path.join(dir_refdb, name_family, underscore_protein_name_refdb)

            os.makedirs(dir_refdb, exist_ok=True)
            os.makedirs(dir_refdb_final, exist_ok=True)

            database = refdb(positive_terms[y], txid[y], name_family, dir_refdb_final, protein_name_refdb, min_len[y], max_len[y])
            logging.info(f"database: {database}")
            make_blast_db(database)

        for y in range(len(tableY)):
            _build_refdb_for_term(y)

        new_tableX = []


        dir_marker = os.path.join(final_dir, "markers")
        os.makedirs(dir_marker, exist_ok=True)

        # ── Main processing loop: sequential or parallel ───────────────────
        if num_threads is not None:
            # ── PARALLEL PATH ─────────────────────────────────────────────
            new_tableX, family_markers = run_parallel_pipeline(
                tableX=tableX,
                tableY=tableY,
                positive_terms_s=positive_terms,
                negative_terms_s=negative_terms,
                min_len_s=min_len,
                max_len_s=max_len,
                txid_s=txid,
                name_fam_s=name_fam,
                ws=ws,
                final_dir=final_dir,
                positive_dict=positive_dict,
                num_workers=num_threads,
            )

        else:
            # ── SEQUENTIAL PATH (original behaviour) ──────────────────────
            for i in range(len(tableX)):
                for j in range(len(tableY)):
                    if tableX.loc[i, 'Family'] == tableY.loc[j, 'Name']:

                        line = dict(tableX.iloc[i])

                        genome_code = line.get("Virus GENBANK accession", "")
                        genus = line.get("Genus", "")
                        Class = line.get("Class", "")
                        family = line.get("Family", "")
                        subfamily = line.get("Subfamily", "")
                        ictv_id = line.get("ICTV_ID", "")

                        if isinstance(family, str):
                            family = family
                        elif isinstance(family, float) and isinstance(subfamily, str):
                            family = subfamily
                        elif isinstance(family, float) and isinstance(subfamily, float):
                            family = 'unclassified'

                        protein_name = positive_terms[j]
                        definitive_protein_name = protein_name

                        if "," in protein_name:
                            main_protein_name = protein_name.split(",")[0].strip()
                        else:
                            main_protein_name = protein_name.strip()

                        underscore_protein_name = _safe_path_name(main_protein_name)

                        logging.info(f"Processing GenBank ID {genome_code} with {definitive_protein_name} as an annotation term")

                        dir_genome = os.path.join(final_dir, "cds_virus_fasta")
                        dir_genome_final = os.path.join(dir_genome, family, genus)

                        dir_refdb_info = os.path.join(final_dir, "refdb")
                        dir_refdb_info_final = os.path.join(dir_refdb_info, family, underscore_protein_name)

                        dir_marker = os.path.join(final_dir, "markers")
                        dir_marker_final = os.path.join(dir_marker, family, underscore_protein_name, "fasta")

                        os.makedirs(dir_marker, exist_ok=True)
                        os.makedirs(dir_marker_final, exist_ok=True)
                        os.makedirs(dir_genome, exist_ok=True)
                        os.makedirs(dir_genome_final, exist_ok=True)

                        ictv_id_url = hyperlink_extrator(ws, "ICTV_ID", ictv_id)

                        # ── Accessions Link: segmented vs simple genome ───────
                        is_segmented = (
                            not pd.isna(genome_code)
                            and genome_code
                            and (":" in str(genome_code) or ";" in str(genome_code))
                        )

                        if is_segmented:
                            genome_code_url = hyperlink_code_seg(genome_code)
                        else:
                            genome_code_url = hyperlink_code(genome_code)

                        Gi = _entrez_cache.get_or_compute(
                            genome_code, 'search',
                            lambda gc=genome_code: search_entrez(gc)
                        )
                        protein = filtered_search(Gi, [positive_terms[j]], [negative_terms[j]], min_len[j], max_len[j])
                        protein_url = hyperlink_Protein_ID(protein)

                        seg = None
                        seg_url = None

                        if pd.isna(genome_code) or not genome_code:
                            pass
                        elif is_segmented:
                            if protein and protein != []:
                                retrive_gen = retrieve_genbank_accession(protein)
                                seg = segment_named(retrive_gen, genome_code)
                                seg_url = hyperlink_segment(seg, genome_code)

                        fasta_file = _entrez_cache.get_or_compute(
                            genome_code, 'cds',
                            lambda gc=genome_code, gi=Gi, d=dir_genome_final: cds_prot(gi, d, gc)
                        )
                        number_prot = _entrez_cache.get_or_compute(
                            genome_code, 'nprot',
                            lambda ff=fasta_file: counter_prot(ff)
                        )
                        fasta_protein = marker_fasta(fasta_file, protein, dir_marker_final, genus, family)
                        family_markers.append(family)

                        dict_key = (str(family).strip(), str(positive_terms[j]).strip())

                        if protein == []:
                            logging.info(f'No protein annotated as {definitive_protein_name} was found on GenBank ID {genome_code}.')
                        else:
                            positive_dict[dict_key][0] += 1
                            logging.info(f'Protein annotated as {definitive_protein_name} was found on GenBank ID: {protein}\n')

                        if protein == []:
                            logging.info(f"Running BLAST similarity search of the {number_prot} protein sequences against a reference database of {definitive_protein_name}")
                            database = refdb(positive_terms[j], txid[j], family, dir_refdb_info_final, protein_name, min_len[j], max_len[j])
                            protein = blast_plus(database, fasta_file)
                            protein_url = hyperlink_Protein_ID(protein)

                            seg = None
                            seg_url = None

                            if pd.isna(genome_code) or not genome_code:
                                pass
                            elif is_segmented:
                                if protein and protein != []:
                                    retrive_gen = retrieve_genbank_accession(protein)
                                    seg = segment_named(retrive_gen, genome_code)
                                    seg_url = hyperlink_segment(seg, genome_code)

                            fasta_protein_s = marker_fasta(fasta_file, protein, dir_marker_final, genus, family)
                            family_markers.append(family)
                            if protein is None:
                                positive_dict[dict_key][2] += 1
                                logging.info(f'No hit found for {definitive_protein_name}.\n')
                            else:
                                positive_dict[dict_key][1] += 1
                                logging.info(f'Positive protein for {definitive_protein_name} found: {protein}.\n')

                        line['min_length'] = min_len[j]
                        line['max_length'] = max_len[j]
                        line['Negative_Terms'] = negative_terms[j]
                        line['Positive_Terms'] = positive_terms[j]
                        line['Protein_ID'] = protein if protein else ""
                        line['Segment'] = seg if seg else ""
                        line['tax_id'] = txid[j]
                        line['ICVT_ID link'] = ictv_id_url
                        line['Accession link(s) per segment(s)'] = genome_code_url
                        line['Protein ID link'] = protein_url if protein_url else ""
                        line['Segment link'] = seg_url if seg_url else ""
                        new_tableX.append(line)

        # ── Redundancy padding ────────────────────────────────────────────
        # Duplicate sequences in small marker FASTAs (1..3 seqs) by doubling
        # until the total exceeds 5 (see DUPLICATES_BY_COUNT) so MAFFT/tabajara
        # can build informative HMMs.
        # redundancy_status is consumed later by the report_hmms builder.
        redundancy_status = {}
        pad_marker_fastas(dir_marker, redundancy_status)

        logging.info('Starting family sequence alignment...')

        for family_folder in os.listdir(dir_marker):
            family_path = os.path.join(dir_marker, family_folder)
            
            if os.path.isdir(family_path):
                logging.info(f'Processing family: {family_folder}')
                
                for marker_folder in os.listdir(family_path):
                    marker_path = os.path.join(family_path, marker_folder)
                    
                    if os.path.isdir(marker_path):
                        logging.info(f'  Processing marker: {marker_folder}')
                        
                        fasta_path = os.path.join(marker_path, "fasta")
                        
                        if os.path.exists(fasta_path):
                            family_fasta_file = os.path.join(fasta_path, f"{family_folder}.fasta")
                            
                            if os.path.exists(family_fasta_file):
                                output_aligned = os.path.join(marker_path, f"{family_folder}_aligned.fasta")
                                run_mafft(family_fasta_file, output_aligned)
                            else:
                                logging.warning(f'    File not found: {family_fasta_file}')
                        else:
                            logging.warning(f'    Fasta folder not found: {fasta_path}')

        logging.info('Sequence alignment completed!\n')

        dir_marker = Path(dir_marker)

        for family_path in dir_marker.iterdir():
            if family_path.is_dir():
                logging.info(f'Processing family: {family_path.name}')

                for marker_path in family_path.iterdir():
                    if marker_path.is_dir():
                        genus_name.clear()
                        logging.info(f'  Processing marker: {marker_path.name}')

                        aligned_file = next(marker_path.glob("*.fasta"), None)

                        if aligned_file is None:
                            logging.info("  .fasta file not found")
                            continue

                        logging.info(f"  Reading file: {aligned_file.name}")

                        with aligned_file.open("r") as f:
                            for line in f:
                                if line.startswith(">"):
                                    name = line.lstrip(">").split("_")[0]
                                    genus_name.append(name)

                        if not genus_name:
                            logging.warning("  Nenhum cabeçalho encontrado no FASTA")
                            continue

                        clean_list = list(dict.fromkeys(genus_name))
                        list_c = ",".join(clean_list)

                        # ── Pass config-driven Tabajara params ────────────
                        run_tabajara_con(aligned_file, list_c, marker_path,
                                         tabajara_params=cfg_tab_con)
                        run_tabajara_dis(aligned_file, list_c, marker_path,
                                         tabajara_params=cfg_tab_dis)

        # ── Collect and rename HMMs ───────────────────────────────────────
        logging.info("Collecting and renaming HMM files...")
        hmm_traceability = collect_and_rename_hmms(dir_marker, final_dir)
        generate_hmm_traceability_report(hmm_traceability, final_dir)

        final_table_csv = f'VMR+_{new_file.name}'
        table_file = os.path.join(final_dir, final_table_csv)

        generate_family_grouped_report(genome_conter, positive_dict, new_tableX)

        data_report_hmms = []

        for family in os.listdir(dir_marker):
            family_path_report = os.path.join(dir_marker, family)

            if not os.path.isdir(family_path_report):
                continue

            markers = [m for m in os.listdir(family_path_report) if os.path.isdir(os.path.join(family_path_report, m))]

            for marker in markers:
                marker_path_report = os.path.join(family_path_report, marker)

                fasta_file_report = os.path.join(marker_path_report, "fasta", f"{family}.fasta")
                seq_count = 0

                if os.path.exists(fasta_file_report):
                    with open(fasta_file_report, "r", errors="ignore") as f_report:
                        seq_count = sum(1 for line_report in f_report if line_report.startswith(">"))

                hmm_path = os.path.join(marker_path_report, "tabajara_family", "hmms", "valid_HMMs")
                hmm_count = 0

                if os.path.exists(hmm_path):
                    hmm_count = len([
                        f_report for f_report in os.listdir(hmm_path)
                        if os.path.isfile(os.path.join(hmm_path, f_report))
                    ])

                fam_redundancy = redundancy_status.get((family, marker, None), False)
                data_report_hmms.append([family, "Family-wide models", marker, seq_count, hmm_count, fam_redundancy])

            for marker in markers:
                marker_path_report = os.path.join(family_path_report, marker)

                genera_base = os.path.join(marker_path_report, "tabajara_genera", "hmms")

                if not os.path.exists(genera_base):
                    continue

                for genus in os.listdir(genera_base):
                    genus_path = os.path.join(genera_base, genus)

                    if not os.path.isdir(genus_path):
                        continue

                    hmm_valid = os.path.join(genus_path, "valid_HMMs")
                    hmm_count = 0

                    if os.path.exists(hmm_valid):
                        hmm_count = len([
                            f_report for f_report in os.listdir(hmm_valid)
                            if os.path.isfile(os.path.join(hmm_valid, f_report))
                        ])

                    fasta_file_report = os.path.join(marker_path_report, "fasta", f"{genus}.fasta")
                    seq_count = 0

                    if os.path.exists(fasta_file_report):
                        with open(fasta_file_report, "r", errors="ignore") as f_report:
                            seq_count = sum(1 for line_report in f_report if line_report.startswith(">"))

                    genus_redundancy = redundancy_status.get((family, marker, genus), False)
                    data_report_hmms.append([family, genus, marker, seq_count, hmm_count, genus_redundancy])

        data_report_hmms.sort(key=lambda x: (x[0], x[1], x[2]))

        report_hmms_csv = f'report_hmms.csv'
        report_hmms_file = os.path.join(final_dir, report_hmms_csv)

        with open(report_hmms_file, "w", newline="") as f_report:
            writer = csv.writer(f_report, delimiter=";")
            writer.writerow(["Family", "Genus", "Marker", "#sequences", "#profile HMMs", "Redundancy_Status"])
            writer.writerows(data_report_hmms)

        report_csv_file = Path(report_hmms_file)
        df_xl_report = pd.read_csv(report_csv_file, delimiter=';')
        xlsx_report_file = report_csv_file.with_suffix('.xlsx')
        df_xl_report.to_excel(xlsx_report_file, index=False)

        logging.info(f"Generated CSV: report_hmms.csv")
        logging.info("Saving all results...\n")
        
        new_tableX = pd.DataFrame(new_tableX)
        new_tableX.to_csv(table_file, sep=';', index=False)

        df_csv = pd.read_csv(table_file, delimiter=';')

        new_order = ['Isolate ID','Species Sort', 'Isolate Sort', 'Realm',
                    'Subrealm', 'Kingdom', 'Subkingdom', 'Phylum', 'Subphylum',
                    'Class', 'Subclass', 'Order', 'Suborder', 'Family', 'Subfamily',
                    'Genus', 'Subgenus', 'Species','ICTV_ID', 'ICVT_ID link',
                    'Exemplar or additional isolate', 'Virus name(s)',
                    'Virus name abbreviation(s)', 'Virus isolate designation',
                    'Virus GENBANK accession', 'tax_id', 'Positive_Terms',
                    'Negative_Terms','min_length','max_length', 'Protein_ID',
                    'Protein ID link','Segment', 'Segment link', 'Genome coverage',
                    'Genome', 'Host source', 'Accessions Link', 'Accession link(s) per segment(s)']

        df_csv = df_csv[new_order]
        df_csv.to_csv(table_file, index=False, sep=';')

        final_table_xl = f'VMR+_{input_file}'
        output_path = os.path.join(final_dir, final_table_xl)
        logging.info('Generating .xlsx file\n')

        csv_file = Path(table_file)
        df_xl_final = pd.read_csv(csv_file, delimiter=';')
        xlsx_file = csv_file.with_suffix('.xlsx')
        df_xl_final.to_excel(xlsx_file, index=False)

        wb_final_table = load_workbook(xlsx_file)
        ws_final_table = wb_final_table.worksheets[0]

        # ── Font style for hyperlink cells ────────────────────────────────
        # Blue (#0563C1) + underline for unvisited; Excel natively renders
        # visited links in purple — no additional code needed for that.
        hyperlink_font = Font(color="0563C1", underline="single")

        # Map: display column name  →  auxiliary column that holds the URL as
        # plain text. The URL is embedded directly INTO the display cell, so
        # the auxiliary column can be deleted afterwards without breaking the
        # hyperlinks.
        hyperlink_column_map = {
            'ICTV_ID':          'ICVT_ID link',
            'Accessions Link':  'Accession link(s) per segment(s)',
            'Protein_ID':       'Protein ID link',
            'Segment':          'Segment link',
        }

        # Build a header -> column-index lookup from the worksheet header row.
        col_index = {}
        for col in range(1, ws_final_table.max_column + 1):
            header = ws_final_table.cell(row=1, column=col).value
            if header is not None:
                col_index[str(header)] = col

        # URLs are read straight from the in-memory DataFrame (df_xl_final) —
        # the reliable source of truth — instead of reading them back from
        # worksheet cells (which was fragile). Each display cell gets an
        # explicit Hyperlink object, which is more robust across openpyxl
        # versions than assigning a bare string. DataFrame row i maps to
        # worksheet row i + 2 (row 1 is the header).
        for display_col, link_col in hyperlink_column_map.items():
            if display_col not in col_index or link_col not in df_xl_final.columns:
                continue  # column absent in this run — skip silently

            d_col = col_index[display_col]

            for df_idx, raw_url in enumerate(df_xl_final[link_col].tolist()):
                if pd.isna(raw_url):
                    continue
                url = str(raw_url).strip()
                if not url or url.lower() == 'nan':
                    continue

                disp_cell = ws_final_table.cell(row=df_idx + 2, column=d_col)
                disp_cell.hyperlink = Hyperlink(ref=disp_cell.coordinate, target=url)
                disp_cell.font = hyperlink_font

        wb_final_table.save(xlsx_file)

        logging.info('Generating final report...\n')
        logging.info('Final report')

        end_time = time.perf_counter()
        total_time = end_time - start_time

        hours, rest = divmod(total_time, 3600)
        minutes, seconds = divmod(rest, 60)

        logging.info(f"Total execution time:{int(hours)} hours, {int(minutes)} minutes e {int(seconds)} seconds")
        logging.info("Execution finished!")