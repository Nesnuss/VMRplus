"""Excel hyperlink and segment-name helpers.

Moved verbatim from the monolithic script (pure move).
"""

import re

import pandas as pd
from openpyxl.utils import get_column_letter


def hyperlink_Protein_ID(protein_id):
    try:
        if protein_id == [] or protein_id is None:
            return 
        else:
            protein_id = str(protein_id).strip()
            hyperlink = f"https://www.ncbi.nlm.nih.gov/protein/{protein_id}"
            return hyperlink
    except Exception:
        return 
    
def hyperlink_code(genome_code):
    try:
        code = str(genome_code).strip()
        hyperlink = f"https://www.ncbi.nlm.nih.gov/nuccore/{code}"
        return hyperlink
    except Exception:
        return

def hyperlink_extrator(ws, col_name, ictv_id):

    code_ictv = str(ictv_id).strip()
    
    data = []  
    
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == col_name:
            letra = get_column_letter(col)
            hyperlink_form_default = re.compile(r'HYPERLINK\("([^"]+)"(?:,\s*"([^"]*)")?\)', re.IGNORECASE)
            
            for cell in ws[letra][1:]:
                if cell.value and code_ictv in str(cell.value):
                    link = None  

                    if cell.hyperlink:
                        link = cell.hyperlink.target
                    elif cell.data_type == 'f' and cell.value:
                        match = hyperlink_form_default.search(cell.value)
                        if match:
                            link = match.group(1)
                    if link:
                        data.append(link)
            break  
    
    if data:
        return data[0]
    else:
        return None


def hyperlink_code_seg(genome_code):

    if not genome_code or genome_code == [] or pd.isna(genome_code) or str(genome_code).strip().lower() == 'nan':
        return 

    genome_code_str = str(genome_code).strip()

    acc_ids = []

    for part in re.split(r';', genome_code_str):
        part = part.strip()
        if not part:
            continue


        if ':' in part:
            _, acc = part.split(':', 1)
            acc = acc.strip()
        else:
            acc = part.strip()

        acc = acc.split('.')[0].strip()

        if (len(acc) >= 4
                and any(c.isalpha() for c in acc)
                and any(c.isdigit() for c in acc)):
            acc_ids.append(acc)

    if not acc_ids:
        return None

    return f"https://www.ncbi.nlm.nih.gov/nuccore/{','.join(acc_ids)}"

def hyperlink_segment(seg_name, genome_code):
    """
    Builds a URL for a named segment of a segmented genome.

    Given the segment label (e.g. 'L') and the full genome_code field
    (e.g. 'L: MK861116; M: MK861117; S: MK861118' or 'MK861116; MK861117'),
    resolves the corresponding GenBank accession ID and returns the NCBI
    nuccore URL.

    Used only for segmented genomes (when genome_code contains ':' or ';').

    Parameters
    ----------
    seg_name    : str — the segment name returned by segment_named (e.g. 'L',
                  'Seg1')
    genome_code : str — the full 'Virus GENBANK accession' field

    Returns
    -------
    str | None  — the NCBI URL, or None if the ID cannot be resolved.
    """
    if not seg_name or not genome_code:
        return None

    genome_code_str = str(genome_code).strip()

    # ── Named format: 'L: MK861116; M: MK861117' ──────────────────────────
    if ':' in genome_code_str:
        for item in genome_code_str.split(';'):
            item = item.strip()
            if not item:
                continue
            try:
                name, acc = item.split(':', 1)
            except ValueError:
                continue
            if name.strip() == seg_name:
                acc_id = acc.strip().split('.')[0]  # remove version suffix
                return f"https://www.ncbi.nlm.nih.gov/nuccore/{acc_id}"

    # ── Unnamed format: positional (Seg1, Seg2, …) ────────────────────────
    # seg_name_named returns 'SegN' (1-based) when no ':' separator is found.
    seg_match = re.match(r'^Seg(\d+)$', str(seg_name))
    if seg_match:
        index = int(seg_match.group(1)) - 1   # convert to 0-based
        ids_list = [
            i.strip() for i in genome_code_str.split(';') if i.strip()
        ]
        if 0 <= index < len(ids_list):
            acc_id = ids_list[index].split('.')[0]
            return f"https://www.ncbi.nlm.nih.gov/nuccore/{acc_id}"

    return None
def segment_named(codeid, syntax):

    if codeid is None:
        return None

    codeid_base = codeid.split(".")[0]

    if ":" in syntax:
        for item in syntax.split(";"):
            item = item.strip()
            if not item:
                continue
            try:
                name, current_id = item.split(":", 1)
            except ValueError:
                continue
            if current_id.strip() == codeid_base:
                return name.strip()
        return None
    else:
        ids_list = [i.strip() for i in syntax.split(";") if i.strip()]
        try:
            index = ids_list.index(codeid_base)
            return f"Seg{index + 1}"
        except ValueError:
            return None

# ──────────────────────────────────────────────────────────────────────────────
# HMM collection, renaming and traceability report
# ──────────────────────────────────────────────────────────────────────────────
