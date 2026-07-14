"""Pipeline orchestration for VMR+: the single-task worker, the parallel
driver and main().

Moved from the monolithic script. The ONLY non-verbatim changes are:
  * the shared singletons live in vmrplus.ncbi, so they are referenced as
    ncbi._ncbi_limiter / ncbi._entrez_cache here;
  * the old ``if __name__ == '__main__'`` block is now def main(args), with
    family_markers/genus_name initialised at its top (previously module-level).
"""

import csv
import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from Bio import Entrez
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from . import ncbi
from .config import (
    detect_cli_config_conflict,
    generate_config_template,
    help_text,
    load_config,
    version,
)
from .external import blast_plus, make_blast_db, run_mafft, run_tabajara_con, run_tabajara_dis
from .hmms import collect_and_rename_hmms, generate_hmm_traceability_report
from .hyperlinks import (
    hyperlink_code,
    hyperlink_code_seg,
    hyperlink_extrator,
    hyperlink_Protein_ID,
    hyperlink_segment,
    segment_named,
)
from .markers import marker_fasta, pad_marker_fastas
from .ncbi import (
    cds_prot,
    download_all_genomes,
    filtered_search,
    refdb,
    retrieve_genbank_accession,
    search_entrez,
)
from .paths import _safe_path_name, counter_prot, unique_dir
from .reporting import generate_family_grouped_report


def _process_single_task(task):
    """
    Executes all network-bound work for one (genome × protein) pair.

    This function is designed to run inside a ThreadPoolExecutor worker.
    It is a self-contained unit: it receives all data it needs via *task*,
    performs the Entrez queries (rate-limited), the refdb download,
    cds_prot download, BLAST search and marker extraction, and returns a
    result dict suitable for appending to new_tableX.

    On unrecoverable failure the function logs the error and returns None
    so the caller can record a partial result and continue.

    Parameters
    ----------
    task : dict with keys
        i, j               — row indices into tableX / tableY
        line               — dict copy of tableX.iloc[i]
        tableY_row         — dict of the matching tableY row
        ws                 — openpyxl worksheet (read-only in threads, safe)
        final_dir          — Path to the output directory
        dir_marker         — str path to markers root dir

    Returns
    -------
    dict | None
    """
    try:
        # ── Unpack task ───────────────────────────────────────────────────
        line              = task['line']
        ty                = task['tableY_row']
        ws                = task['ws']
        final_dir         = task['final_dir']

        genome_code       = line.get("Virus GENBANK accession", "")
        genus             = line.get("Genus", "")
        family            = line.get("Family", "")
        subfamily         = line.get("Subfamily", "")
        ictv_id           = line.get("ICTV_ID", "")

        if isinstance(family, str):
            pass
        elif isinstance(family, float) and isinstance(subfamily, str):
            family = subfamily
        elif isinstance(family, float) and isinstance(subfamily, float):
            family = 'unclassified'

        protein_name          = ty['positive_terms']
        definitive_prot_name  = protein_name
        negative_terms_val    = ty['negative_terms']
        min_len_val           = ty['min_len']
        max_len_val           = ty['max_len']
        txid_val              = ty['txid']

        if "," in protein_name:
            main_protein_name = protein_name.split(",")[0].strip()
        else:
            main_protein_name = protein_name.strip()

        underscore_protein_name = _safe_path_name(main_protein_name)

        logging.info(
            f"[parallel] Processing GenBank ID {genome_code} "
            f"with {definitive_prot_name} as annotation term"
        )

        # ── Build sub-directories (thread-safe: exist_ok=True) ───────────
        dir_genome       = os.path.join(final_dir, "cds_virus_fasta")
        dir_genome_final = os.path.join(dir_genome, family, genus)
        dir_refdb_info   = os.path.join(final_dir, "refdb")
        dir_refdb_final  = os.path.join(dir_refdb_info, family, underscore_protein_name)
        dir_marker       = os.path.join(final_dir, "markers")
        dir_marker_final = os.path.join(dir_marker, family, underscore_protein_name, "fasta")

        os.makedirs(dir_genome_final,  exist_ok=True)
        os.makedirs(dir_refdb_final,   exist_ok=True)
        os.makedirs(dir_marker_final,  exist_ok=True)

        # ── Hyperlinks (read-only worksheet access) ───────────────────────
        ictv_id_url = hyperlink_extrator(ws, "ICTV_ID", ictv_id)

        genome_code_valido = (
            genome_code
            and not pd.isna(genome_code)
            and str(genome_code).strip().lower() != 'nan'
            and str(genome_code).strip() != ''
        )

        if not genome_code_valido:
            genome_code_url = None
        else:
            is_segmented = (
                ":" in str(genome_code) or ";" in str(genome_code)
            )
            if is_segmented:
                genome_code_url = hyperlink_code_seg(genome_code)
            else:
                genome_code_url = hyperlink_code(genome_code)

        # ── Cached / rate-limited Entrez calls ────────────────────────────
        # search_entrez and cds_prot depend only on genome_code; when
        # multiple markers are processed for the same genome the results
        # are reused via _entrez_cache, avoiding redundant NCBI calls.
        Gi = ncbi._entrez_cache.get_or_compute(
            genome_code, 'search',
            lambda: search_entrez(genome_code)
        )
        protein = filtered_search(
            Gi, [protein_name], [negative_terms_val], min_len_val, max_len_val
        )
        protein_url = hyperlink_Protein_ID(protein)

        seg     = None
        seg_url = None

        if not (pd.isna(genome_code) or not genome_code) and is_segmented:
            if protein and protein != []:
                retrive_gen = retrieve_genbank_accession(protein)
                seg     = segment_named(retrive_gen, genome_code)
                seg_url = hyperlink_segment(seg, genome_code)

        fasta_file = ncbi._entrez_cache.get_or_compute(
            genome_code, 'cds',
            lambda: cds_prot(Gi, dir_genome_final, genome_code)
        )
        number_prot = ncbi._entrez_cache.get_or_compute(
            genome_code, 'nprot',
            lambda: counter_prot(fasta_file)
        )
        marker_fasta(fasta_file, protein, dir_marker_final, genus, family)

        # ── Annotation counters (returned; aggregated by caller) ──────────
        annotation_hit  = 0
        similarity_hit  = 0
        undetected      = 0

        if protein == []:
            logging.info(
                f'[parallel] No protein annotated as {definitive_prot_name} '
                f'found on GenBank ID {genome_code}.'
            )
        else:
            annotation_hit = 1
            logging.info(
                f'[parallel] Protein annotated as {definitive_prot_name} '
                f'found: {protein}'
            )

        # ── BLAST fallback ────────────────────────────────────────────────
        if protein == []:
            logging.info(
                f'[parallel] Running BLAST on {number_prot} proteins '
                f'vs reference DB for {definitive_prot_name}'
            )
            database = refdb(
                protein_name, txid_val, family,
                dir_refdb_final, protein_name, min_len_val, max_len_val
            )
            protein     = blast_plus(database, fasta_file)
            protein_url = hyperlink_Protein_ID(protein)

            seg     = None
            seg_url = None

            if not (pd.isna(genome_code) or not genome_code) and is_segmented:
                if protein and protein != []:
                    retrive_gen = retrieve_genbank_accession(protein)
                    seg     = segment_named(retrive_gen, genome_code)
                    seg_url = hyperlink_segment(seg, genome_code)

            marker_fasta(fasta_file, protein, dir_marker_final, genus, family)

            if protein is None:
                undetected = 1
                logging.info(f'[parallel] No hit found for {definitive_prot_name}.')
            else:
                similarity_hit = 1
                logging.info(
                    f'[parallel] Positive protein for {definitive_prot_name}: {protein}.'
                )

        # ── Assemble result row ───────────────────────────────────────────
        result_line = dict(line)
        result_line['min_length']        = min_len_val
        result_line['max_length']        = max_len_val
        result_line['Negative_Terms']    = negative_terms_val
        result_line['Positive_Terms']    = protein_name
        result_line['Protein_ID']        = protein if protein else ""
        result_line['Segment']           = seg if seg else ""
        result_line['tax_id']            = txid_val
        result_line['ICVT_ID link']      = ictv_id_url
        result_line['Accession link(s) per segment(s)'] = genome_code_url
        result_line['Protein ID link']   = protein_url if protein_url else ""
        result_line['Segment link']      = seg_url if seg_url else ""

        return {
            'row':            result_line,
            'family':         str(family).strip(),
            'positive_terms': str(protein_name).strip(),
            'annotation':     annotation_hit,
            'similarity':     similarity_hit,
            'undetected':     undetected,
        }

    except Exception as exc:
        logging.error(
            f"[parallel] Unrecoverable error for genome_code="
            f"{task.get('line', {}).get('Virus GENBANK accession', '?')} "
            f"/ protein={task.get('tableY_row', {}).get('positive_terms', '?')}: "
            f"{exc}",
            exc_info=True,
        )
        return None   # caller will skip and continue




def run_parallel_pipeline(
    tableX, tableY,
    positive_terms_s, negative_terms_s,
    min_len_s, max_len_s, txid_s, name_fam_s,
    ws, final_dir, positive_dict,
    num_workers,
):
    """
    Parallel entry point that replaces the sequential double for-loop over
    tableX × tableY.

    Builds one task per matching (i, j) pair, submits them to a
    ThreadPoolExecutor, collects results in original order, and returns
    (new_tableX, family_markers_list).

    Thread-safety notes
    -------------------
    * os.makedirs(..., exist_ok=True)  — safe for concurrent calls.
    * openpyxl worksheet (ws) — used read-only; safe.
    * positive_dict — updated by the *caller* after all futures resolve,
      so no concurrent writes occur.
    * logging — Python's logging module is thread-safe.
    * File writes (cds_prot, marker_fasta) — each thread writes to a
      unique path derived from genome_code + genus, so no collision.
    """

    # ── Build task list (preserves original order) ────────────────────────
    tasks = []
    for i in range(len(tableX)):
        for j in range(len(tableY)):
            if tableX.loc[i, 'Family'] != tableY.loc[j, 'Name']:
                continue
            tasks.append({
                'i':    i,
                'j':    j,
                'line': dict(tableX.iloc[i]),
                'tableY_row': {
                    'positive_terms': positive_terms_s[j],
                    'negative_terms': negative_terms_s[j],
                    'min_len':        min_len_s[j],
                    'max_len':        max_len_s[j],
                    'txid':           txid_s[j],
                    'name_fam':       name_fam_s[j],
                },
                'ws':        ws,
                'final_dir': final_dir,
            })

    logging.info(
        f"[parallel] Submitting {len(tasks)} tasks to "
        f"{num_workers} worker threads."
    )

    new_tableX      = []
    family_markers  = []

    # ── Submit and collect in order ───────────────────────────────────────
    # We use a list of futures indexed by task position so that the output
    # table preserves the same row order as the sequential version.
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        future_to_task = {
            executor.submit(_process_single_task, t): t
            for t in tasks
        }

        completed = 0
        for future in as_completed(future_to_task):
            completed += 1
            t = future_to_task[future]
            genome_label = t['line'].get('Virus GENBANK accession', '?')
            prot_label   = t['tableY_row']['positive_terms']

            try:
                result = future.result()
            except Exception as exc:
                logging.error(
                    f"[parallel] Future raised for genome={genome_label} "
                    f"protein={prot_label}: {exc}",
                    exc_info=True,
                )
                result = None

            if result is None:
                # Partial failure: log and skip this row
                logging.warning(
                    f"[parallel] Skipping row genome={genome_label} "
                    f"protein={prot_label} due to unrecoverable error."
                )
                continue

            # ── Aggregate counters ────────────────────────────────────────
            dict_key = (result['family'], result['positive_terms'])
            if dict_key in positive_dict:
                positive_dict[dict_key][0] += result['annotation']
                positive_dict[dict_key][1] += result['similarity']
                positive_dict[dict_key][2] += result['undetected']

            new_tableX.append(result['row'])
            family_markers.append(result['family'])

            logging.info(
                f"[parallel] Completed {completed}/{len(tasks)} — "
                f"genome={genome_label}"
            )

    # Sort results back to original tableX order (as_completed is unordered)
    # We use the original (i, j) indices preserved in the task dict to
    # reconstruct a stable sort key.
    task_order = {
        (t['i'], t['j']): idx for idx, t in enumerate(tasks)
    }

    def _sort_key(row):
        # Match row back to task by genome_code + Positive_Terms
        gc = row.get('Virus GENBANK accession', '')
        pt = row.get('Positive_Terms', '')
        for (i, j), idx in task_order.items():
            t = tasks[idx]
            if (t['line'].get('Virus GENBANK accession', '') == gc
                    and t['tableY_row']['positive_terms'] == pt):
                return idx
        return 999999

    new_tableX.sort(key=_sort_key)

    logging.info(
        f"[parallel] Pipeline complete. "
        f"{len(new_tableX)}/{len(tasks)} rows collected."
    )

    return new_tableX, family_markers


# ──────────────────────────────────────────────────────────────────────────────
# Argument parsing
# ──────────────────────────────────────────────────────────────────────────────



def main(args):
    family_markers = []
    genus_name = []

    # ── --generate-config: generate template and exit immediately ─────────
    if args.generate_config:
        generate_config_template()
        sys.exit(0)

    if not len(sys.argv) > 1:
        print(help_text)

    elif args.h:
        print(help_text)

    elif args.v:
        print(f"""
VMR Program version {version} - 30 jan 2026
(c) 2024. Rafael Santos da Silva & Arthur Gruber
""")

    else:

        # ── Conflict detection ────────────────────────────────────────────
        detect_cli_config_conflict(sys.argv[1:], config_flag_used=args.config is not None)

        # ── Load config (if supplied) or fall back to CLI args ────────────
        cfg_general     = {}
        cfg_tab_con     = None   # None → run_tabajara_con uses its own defaults
        cfg_tab_dis     = None   # None → run_tabajara_dis uses its own defaults

        if args.config:
            raw_cfg = load_config(args.config)

            cfg_general    = raw_cfg.get('general', {})
            cfg_tab_con    = raw_cfg.get('tabajara_con')
            cfg_tab_dis    = raw_cfg.get('tabajara_dis')

            # Resolve final parameter values from config
            input_file  = cfg_general.get('input') or None
            output_dir  = cfg_general.get('output', 'output_dir')
            sheet_num   = int(cfg_general.get('sheet', 1))
            term_file   = cfg_general.get('terms') or None
            term_sheet  = int(cfg_general.get('terms_sheet', 1))
            email_val   = cfg_general.get('email', Entrez.email)
            api_key_val = cfg_general.get('api_key', Entrez.api_key)
            gb_val      = cfg_general.get('gb', args.gb)

            if not input_file:
                print("Error: 'input' is required. Set it in the config file or use -i.")
                sys.exit(1)
            if not term_file:
                print("Error: 'terms' is required. Set it in the config file or use -t.")
                sys.exit(1)

        else:
            # Traditional CLI path
            input_file  = args.i
            output_dir  = args.o
            sheet_num   = args.s
            term_file   = args.t
            term_sheet  = args.ts
            email_val   = Entrez.email
            api_key_val = Entrez.api_key
            gb_val      = args.gb

            if not input_file:
                print("Error: input file is required. Use -i <file> or -c <config.ini>.")
                sys.exit(1)
            if not term_file:
                print("Error: terms file is required. Use -t <file> or -c <config.ini>.")
                sys.exit(1)

        # Whether to download all genomes (-gb yes / gb = yes in config)
        download_genomes = str(gb_val).strip().lower() == 'yes' if gb_val else False

        # Apply Entrez credentials (may have been overridden by config)
        Entrez.email   = email_val
        Entrez.api_key = api_key_val

        # ── Output directory ──────────────────────────────────────────────
        final_dir = unique_dir(output_dir)
        os.makedirs(final_dir)

        log_file = os.path.join(final_dir, "VMR.log")

        logging.basicConfig(
            filename=log_file,
            filemode='a',
            format='%(asctime)s - %(levelname)s - %(message)s',
            level=logging.INFO
        )

        # Emit deferred warnings for unknown [general] keys
        for unknown_key in raw_cfg.get('_unknown_general', []) if args.config else []:
            logging.warning(
                f"Config file: unknown key '{unknown_key}' in [general] section -- ignored."
            )

        # ── Parallel mode setup ───────────────────────────────────────────
        num_threads = args.threads   # None = sequential

        if num_threads is not None:
            if num_threads < 1:
                print("Error: -thread/-thr must be a positive integer (>= 1).")
                sys.exit(1)
            if num_threads > 10:
                logging.warning(
                    f"-thread value {num_threads} exceeds the NCBI API limit of "
                    f"10 req/s. Consider using -thread 10 or lower to avoid "
                    f"HTTP 429 errors."
                )
            if not Entrez.api_key:
                logging.warning(
                    "Parallel mode (-thread) is active but no NCBI API key is set. "
                    "Without an API key the NCBI limit is 3 req/s. "
                    "Set api_key in the config file or hardcode Entrez.api_key."
                )
            # Rate: with an API key NCBI's hard ceiling is 10 req/s.  The
            # token-bucket limiter below enforces that ceiling exactly, so
            # we no longer need to shave off extra headroom by capping at 8;
            # we just make sure not to exceed 10 req/s regardless of how
            # many threads are requested.
            max_rate = min(num_threads, 10)   # cap at the NCBI ceiling (10 req/s)
            ncbi._ncbi_limiter = ncbi.NCBIRateLimiter(max_rate=max_rate)
            logging.info(
                f"Parallel mode enabled: {num_threads} worker thread(s), "
                f"rate limiter set to {max_rate} req/s."
            )
        else:
            logging.info("Sequential mode (use -thread N to enable parallelism).")

        start_time = time.perf_counter()
        logging.info("Starting execution...")
        if args.config:
            logging.info(f"Configuration loaded from: {args.config}")

        # ── VMR/MSL table → CSV ───────────────────────────────────────────
        sheet = sheet_num - 1
        wb = load_workbook(input_file)
        ws = wb.worksheets[sheet]

        df_xl = pd.read_excel(input_file, sheet_name=sheet)
        file = Path(input_file)
        new_file = file.with_suffix(".csv")
        csv_file = os.path.join(final_dir, new_file.name)
        df_xl.to_csv(csv_file, sep=";", index=False)
        logging.info(f"{file.name} to {new_file.name} conversion")

        # ── Terms table → CSV ─────────────────────────────────────────────
        term_sheet_idx = term_sheet - 1
        term_df_xl = pd.read_excel(term_file, sheet_name=term_sheet_idx)
        tfile = Path(term_file)
        term_new_file = tfile.with_suffix(".csv")
        term_csv_file = os.path.join(final_dir, term_new_file.name)
        term_df_xl.to_csv(term_csv_file, sep=";", index=False)
        logging.info(f"{tfile.name} to {term_new_file.name} conversion")

        tableX = pd.DataFrame(pd.read_csv(csv_file, delimiter=';'))
        tableY = pd.read_csv(term_csv_file, delimiter=';')
        genome_conter = tableX['Virus GENBANK accession'].nunique()

        txid = tableY['tax_id'] 
        positive_terms = tableY['Positive_terms']
        negative_terms = tableY['Negative_terms']
        min_len = tableY['min_length']
        max_len = tableY['max_length']
        parent = tableY['Parent']
        name_fam = tableY["Name"]

        positive_dict = {}
        for y_init in range(len(tableY)):
            key = (str(name_fam[y_init]).strip(), str(positive_terms[y_init]).strip())
            positive_dict[key] = [0, 0, 0]

        # ── Entrez results cache (avoids duplicate NCBI calls) ────────────
        ncbi._entrez_cache = ncbi.EntrezCache()
        logging.info("Entrez results cache initialised.")

        # ── Optional whole-genome download step (-gb yes) ─────────────────
        # Runs before the marker pipeline; the normal pipeline continues
        # afterwards regardless.
        if download_genomes:
            download_all_genomes(tableX, final_dir, num_threads)

        def _build_refdb_for_term(y):
            """
            Builds the reference protein database (and its BLAST index) for
            a single row of the terms table.
            """
            name_family = name_fam[y]
            protein_name_refdb = positive_terms[y]

            if pd.isna(protein_name_refdb) or not str(protein_name_refdb).strip():
                logging.warning(
                    f"Row {y} in terms table has an empty 'Positive_terms' "
                    f"value (Name='{name_family}'). Skipping refdb build for this row."
                )
                return

            protein_name_refdb = str(protein_name_refdb)

            if "," in protein_name_refdb:
                main_protein_name_refdb = protein_name_refdb.split(",")[0].strip()
            else:
                main_protein_name_refdb = protein_name_refdb.strip()

            underscore_protein_name_refdb = _safe_path_name(main_protein_name_refdb)

            dir_refdb = os.path.join(final_dir, "refdb")
            dir_refdb_final = os.path.join(dir_refdb, name_family, underscore_protein_name_refdb)

            os.makedirs(dir_refdb, exist_ok=True)
            os.makedirs(dir_refdb_final, exist_ok=True)

            database = refdb(positive_terms[y], txid[y], name_family, dir_refdb_final, protein_name_refdb, min_len[y], max_len[y])
            logging.info(f"database: {database}")
            make_blast_db(database)

        for y in range(len(tableY)):
            _build_refdb_for_term(y)

        new_tableX = []


        dir_marker = os.path.join(final_dir, "markers")
        os.makedirs(dir_marker, exist_ok=True)

        # ── Main processing loop: sequential or parallel ───────────────────
        if num_threads is not None:
            # ── PARALLEL PATH ─────────────────────────────────────────────
            new_tableX, family_markers = run_parallel_pipeline(
                tableX=tableX,
                tableY=tableY,
                positive_terms_s=positive_terms,
                negative_terms_s=negative_terms,
                min_len_s=min_len,
                max_len_s=max_len,
                txid_s=txid,
                name_fam_s=name_fam,
                ws=ws,
                final_dir=final_dir,
                positive_dict=positive_dict,
                num_workers=num_threads,
            )

        else:
            # ── SEQUENTIAL PATH (original behaviour) ──────────────────────
            for i in range(len(tableX)):
                for j in range(len(tableY)):
                    if tableX.loc[i, 'Family'] == tableY.loc[j, 'Name']:

                        line = dict(tableX.iloc[i])

                        genome_code = line.get("Virus GENBANK accession", "")
                        genus = line.get("Genus", "")
                        Class = line.get("Class", "")
                        family = line.get("Family", "")
                        subfamily = line.get("Subfamily", "")
                        ictv_id = line.get("ICTV_ID", "")

                        if isinstance(family, str):
                            family = family
                        elif isinstance(family, float) and isinstance(subfamily, str):
                            family = subfamily
                        elif isinstance(family, float) and isinstance(subfamily, float):
                            family = 'unclassified'

                        protein_name = positive_terms[j]
                        definitive_protein_name = protein_name

                        if "," in protein_name:
                            main_protein_name = protein_name.split(",")[0].strip()
                        else:
                            main_protein_name = protein_name.strip()

                        underscore_protein_name = _safe_path_name(main_protein_name)

                        logging.info(f"Processing GenBank ID {genome_code} with {definitive_protein_name} as an annotation term")

                        dir_genome = os.path.join(final_dir, "cds_virus_fasta")
                        dir_genome_final = os.path.join(dir_genome, family, genus)

                        dir_refdb_info = os.path.join(final_dir, "refdb")
                        dir_refdb_info_final = os.path.join(dir_refdb_info, family, underscore_protein_name)

                        dir_marker = os.path.join(final_dir, "markers")
                        dir_marker_final = os.path.join(dir_marker, family, underscore_protein_name, "fasta")

                        os.makedirs(dir_marker, exist_ok=True)
                        os.makedirs(dir_marker_final, exist_ok=True)
                        os.makedirs(dir_genome, exist_ok=True)
                        os.makedirs(dir_genome_final, exist_ok=True)

                        ictv_id_url = hyperlink_extrator(ws, "ICTV_ID", ictv_id)

                        # ── Accessions Link: segmented vs simple genome ───────
                        is_segmented = (
                            not pd.isna(genome_code)
                            and genome_code
                            and (":" in str(genome_code) or ";" in str(genome_code))
                        )

                        if is_segmented:
                            genome_code_url = hyperlink_code_seg(genome_code)
                        else:
                            genome_code_url = hyperlink_code(genome_code)

                        Gi = ncbi._entrez_cache.get_or_compute(
                            genome_code, 'search',
                            lambda gc=genome_code: search_entrez(gc)
                        )
                        protein = filtered_search(Gi, [positive_terms[j]], [negative_terms[j]], min_len[j], max_len[j])
                        protein_url = hyperlink_Protein_ID(protein)

                        seg = None
                        seg_url = None

                        if pd.isna(genome_code) or not genome_code:
                            pass
                        elif is_segmented:
                            if protein and protein != []:
                                retrive_gen = retrieve_genbank_accession(protein)
                                seg = segment_named(retrive_gen, genome_code)
                                seg_url = hyperlink_segment(seg, genome_code)

                        fasta_file = ncbi._entrez_cache.get_or_compute(
                            genome_code, 'cds',
                            lambda gc=genome_code, gi=Gi, d=dir_genome_final: cds_prot(gi, d, gc)
                        )
                        number_prot = ncbi._entrez_cache.get_or_compute(
                            genome_code, 'nprot',
                            lambda ff=fasta_file: counter_prot(ff)
                        )
                        fasta_protein = marker_fasta(fasta_file, protein, dir_marker_final, genus, family)
                        family_markers.append(family)

                        dict_key = (str(family).strip(), str(positive_terms[j]).strip())

                        if protein == []:
                            logging.info(f'No protein annotated as {definitive_protein_name} was found on GenBank ID {genome_code}.')
                        else:
                            positive_dict[dict_key][0] += 1
                            logging.info(f'Protein annotated as {definitive_protein_name} was found on GenBank ID: {protein}\n')

                        if protein == []:
                            logging.info(f"Running BLAST similarity search of the {number_prot} protein sequences against a reference database of {definitive_protein_name}")
                            database = refdb(positive_terms[j], txid[j], family, dir_refdb_info_final, protein_name, min_len[j], max_len[j])
                            protein = blast_plus(database, fasta_file)
                            protein_url = hyperlink_Protein_ID(protein)

                            seg = None
                            seg_url = None

                            if pd.isna(genome_code) or not genome_code:
                                pass
                            elif is_segmented:
                                if protein and protein != []:
                                    retrive_gen = retrieve_genbank_accession(protein)
                                    seg = segment_named(retrive_gen, genome_code)
                                    seg_url = hyperlink_segment(seg, genome_code)

                            fasta_protein_s = marker_fasta(fasta_file, protein, dir_marker_final, genus, family)
                            family_markers.append(family)
                            if protein is None:
                                positive_dict[dict_key][2] += 1
                                logging.info(f'No hit found for {definitive_protein_name}.\n')
                            else:
                                positive_dict[dict_key][1] += 1
                                logging.info(f'Positive protein for {definitive_protein_name} found: {protein}.\n')

                        line['min_length'] = min_len[j]
                        line['max_length'] = max_len[j]
                        line['Negative_Terms'] = negative_terms[j]
                        line['Positive_Terms'] = positive_terms[j]
                        line['Protein_ID'] = protein if protein else ""
                        line['Segment'] = seg if seg else ""
                        line['tax_id'] = txid[j]
                        line['ICVT_ID link'] = ictv_id_url
                        line['Accession link(s) per segment(s)'] = genome_code_url
                        line['Protein ID link'] = protein_url if protein_url else ""
                        line['Segment link'] = seg_url if seg_url else ""
                        new_tableX.append(line)

        # ── Redundancy padding ────────────────────────────────────────────
        # Duplicate sequences in small marker FASTAs (1..3 seqs) by doubling
        # until the total exceeds 5 (see DUPLICATES_BY_COUNT) so MAFFT/tabajara
        # can build informative HMMs.
        # redundancy_status is consumed later by the report_hmms builder.
        redundancy_status = {}
        pad_marker_fastas(dir_marker, redundancy_status)

        logging.info('Starting family sequence alignment...')

        for family_folder in os.listdir(dir_marker):
            family_path = os.path.join(dir_marker, family_folder)
            
            if os.path.isdir(family_path):
                logging.info(f'Processing family: {family_folder}')
                
                for marker_folder in os.listdir(family_path):
                    marker_path = os.path.join(family_path, marker_folder)
                    
                    if os.path.isdir(marker_path):
                        logging.info(f'  Processing marker: {marker_folder}')
                        
                        fasta_path = os.path.join(marker_path, "fasta")
                        
                        if os.path.exists(fasta_path):
                            family_fasta_file = os.path.join(fasta_path, f"{family_folder}.fasta")
                            
                            if os.path.exists(family_fasta_file):
                                output_aligned = os.path.join(marker_path, f"{family_folder}_aligned.fasta")
                                run_mafft(family_fasta_file, output_aligned)
                            else:
                                logging.warning(f'    File not found: {family_fasta_file}')
                        else:
                            logging.warning(f'    Fasta folder not found: {fasta_path}')

        logging.info('Sequence alignment completed!\n')

        dir_marker = Path(dir_marker)

        for family_path in dir_marker.iterdir():
            if family_path.is_dir():
                logging.info(f'Processing family: {family_path.name}')

                for marker_path in family_path.iterdir():
                    if marker_path.is_dir():
                        genus_name.clear()
                        logging.info(f'  Processing marker: {marker_path.name}')

                        aligned_file = next(marker_path.glob("*.fasta"), None)

                        if aligned_file is None:
                            logging.info("  .fasta file not found")
                            continue

                        logging.info(f"  Reading file: {aligned_file.name}")

                        with aligned_file.open("r") as f:
                            for line in f:
                                if line.startswith(">"):
                                    name = line.lstrip(">").split("_")[0]
                                    genus_name.append(name)

                        if not genus_name:
                            logging.warning("  Nenhum cabeçalho encontrado no FASTA")
                            continue

                        clean_list = list(dict.fromkeys(genus_name))
                        list_c = ",".join(clean_list)

                        # ── Pass config-driven Tabajara params ────────────
                        run_tabajara_con(aligned_file, list_c, marker_path,
                                         tabajara_params=cfg_tab_con)
                        run_tabajara_dis(aligned_file, list_c, marker_path,
                                         tabajara_params=cfg_tab_dis)

        # ── Collect and rename HMMs ───────────────────────────────────────
        logging.info("Collecting and renaming HMM files...")
        hmm_traceability = collect_and_rename_hmms(dir_marker, final_dir)
        generate_hmm_traceability_report(hmm_traceability, final_dir)

        final_table_csv = f'VMR+_{new_file.name}'
        table_file = os.path.join(final_dir, final_table_csv)

        generate_family_grouped_report(genome_conter, positive_dict, new_tableX)

        data_report_hmms = []

        for family in os.listdir(dir_marker):
            family_path_report = os.path.join(dir_marker, family)

            if not os.path.isdir(family_path_report):
                continue

            markers = [m for m in os.listdir(family_path_report) if os.path.isdir(os.path.join(family_path_report, m))]

            for marker in markers:
                marker_path_report = os.path.join(family_path_report, marker)

                fasta_file_report = os.path.join(marker_path_report, "fasta", f"{family}.fasta")
                seq_count = 0

                if os.path.exists(fasta_file_report):
                    with open(fasta_file_report, "r", errors="ignore") as f_report:
                        seq_count = sum(1 for line_report in f_report if line_report.startswith(">"))

                hmm_path = os.path.join(marker_path_report, "tabajara_family", "hmms", "valid_HMMs")
                hmm_count = 0

                if os.path.exists(hmm_path):
                    hmm_count = len([
                        f_report for f_report in os.listdir(hmm_path)
                        if os.path.isfile(os.path.join(hmm_path, f_report))
                    ])

                fam_redundancy = redundancy_status.get((family, marker, None), False)
                data_report_hmms.append([family, "Family-wide models", marker, seq_count, hmm_count, fam_redundancy])

            for marker in markers:
                marker_path_report = os.path.join(family_path_report, marker)

                genera_base = os.path.join(marker_path_report, "tabajara_genera", "hmms")

                if not os.path.exists(genera_base):
                    continue

                for genus in os.listdir(genera_base):
                    genus_path = os.path.join(genera_base, genus)

                    if not os.path.isdir(genus_path):
                        continue

                    hmm_valid = os.path.join(genus_path, "valid_HMMs")
                    hmm_count = 0

                    if os.path.exists(hmm_valid):
                        hmm_count = len([
                            f_report for f_report in os.listdir(hmm_valid)
                            if os.path.isfile(os.path.join(hmm_valid, f_report))
                        ])

                    fasta_file_report = os.path.join(marker_path_report, "fasta", f"{genus}.fasta")
                    seq_count = 0

                    if os.path.exists(fasta_file_report):
                        with open(fasta_file_report, "r", errors="ignore") as f_report:
                            seq_count = sum(1 for line_report in f_report if line_report.startswith(">"))

                    genus_redundancy = redundancy_status.get((family, marker, genus), False)
                    data_report_hmms.append([family, genus, marker, seq_count, hmm_count, genus_redundancy])

        data_report_hmms.sort(key=lambda x: (x[0], x[1], x[2]))

        report_hmms_csv = 'report_hmms.csv'
        report_hmms_file = os.path.join(final_dir, report_hmms_csv)

        with open(report_hmms_file, "w", newline="") as f_report:
            writer = csv.writer(f_report, delimiter=";")
            writer.writerow(["Family", "Genus", "Marker", "#sequences", "#profile HMMs", "Redundancy_Status"])
            writer.writerows(data_report_hmms)

        report_csv_file = Path(report_hmms_file)
        df_xl_report = pd.read_csv(report_csv_file, delimiter=';')
        xlsx_report_file = report_csv_file.with_suffix('.xlsx')
        df_xl_report.to_excel(xlsx_report_file, index=False)

        logging.info("Generated CSV: report_hmms.csv")
        logging.info("Saving all results...\n")
        
        new_tableX = pd.DataFrame(new_tableX)
        new_tableX.to_csv(table_file, sep=';', index=False)

        df_csv = pd.read_csv(table_file, delimiter=';')

        new_order = ['Isolate ID','Species Sort', 'Isolate Sort', 'Realm',
                    'Subrealm', 'Kingdom', 'Subkingdom', 'Phylum', 'Subphylum',
                    'Class', 'Subclass', 'Order', 'Suborder', 'Family', 'Subfamily',
                    'Genus', 'Subgenus', 'Species','ICTV_ID', 'ICVT_ID link',
                    'Exemplar or additional isolate', 'Virus name(s)',
                    'Virus name abbreviation(s)', 'Virus isolate designation',
                    'Virus GENBANK accession', 'tax_id', 'Positive_Terms',
                    'Negative_Terms','min_length','max_length', 'Protein_ID',
                    'Protein ID link','Segment', 'Segment link', 'Genome coverage',
                    'Genome', 'Host source', 'Accessions Link', 'Accession link(s) per segment(s)']

        df_csv = df_csv[new_order]
        df_csv.to_csv(table_file, index=False, sep=';')

        final_table_xl = f'VMR+_{input_file}'
        output_path = os.path.join(final_dir, final_table_xl)
        logging.info('Generating .xlsx file\n')

        csv_file = Path(table_file)
        df_xl_final = pd.read_csv(csv_file, delimiter=';')
        xlsx_file = csv_file.with_suffix('.xlsx')
        df_xl_final.to_excel(xlsx_file, index=False)

        wb_final_table = load_workbook(xlsx_file)
        ws_final_table = wb_final_table.worksheets[0]

        # ── Font style for hyperlink cells ────────────────────────────────
        # Blue (#0563C1) + underline for unvisited; Excel natively renders
        # visited links in purple — no additional code needed for that.
        hyperlink_font = Font(color="0563C1", underline="single")

        # Map: display column name  →  auxiliary column that holds the URL as
        # plain text. The URL is embedded directly INTO the display cell, so
        # the auxiliary column can be deleted afterwards without breaking the
        # hyperlinks.
        hyperlink_column_map = {
            'ICTV_ID':          'ICVT_ID link',
            'Accessions Link':  'Accession link(s) per segment(s)',
            'Protein_ID':       'Protein ID link',
            'Segment':          'Segment link',
        }

        # Build a header -> column-index lookup from the worksheet header row.
        col_index = {}
        for col in range(1, ws_final_table.max_column + 1):
            header = ws_final_table.cell(row=1, column=col).value
            if header is not None:
                col_index[str(header)] = col

        # URLs are read straight from the in-memory DataFrame (df_xl_final) —
        # the reliable source of truth — instead of reading them back from
        # worksheet cells (which was fragile). Each display cell gets an
        # explicit Hyperlink object, which is more robust across openpyxl
        # versions than assigning a bare string. DataFrame row i maps to
        # worksheet row i + 2 (row 1 is the header).
        for display_col, link_col in hyperlink_column_map.items():
            if display_col not in col_index or link_col not in df_xl_final.columns:
                continue  # column absent in this run — skip silently

            d_col = col_index[display_col]

            for df_idx, raw_url in enumerate(df_xl_final[link_col].tolist()):
                if pd.isna(raw_url):
                    continue
                url = str(raw_url).strip()
                if not url or url.lower() == 'nan':
                    continue

                disp_cell = ws_final_table.cell(row=df_idx + 2, column=d_col)
                disp_cell.hyperlink = Hyperlink(ref=disp_cell.coordinate, target=url)
                disp_cell.font = hyperlink_font

        wb_final_table.save(xlsx_file)

        logging.info('Generating final report...\n')
        logging.info('Final report')

        end_time = time.perf_counter()
        total_time = end_time - start_time

        hours, rest = divmod(total_time, 3600)
        minutes, seconds = divmod(rest, 60)

        logging.info(f"Total execution time:{int(hours)} hours, {int(minutes)} minutes e {int(seconds)} seconds")
